import io
import re
from datetime import datetime, timedelta
from app.db import get_db


# ---------------------------------------------------------------------------
# Sort helpers – Croatian collation + natural (numeric-aware) sort
# ---------------------------------------------------------------------------

_HR_MAP = str.maketrans({
    'č': 'c\x01', 'Č': 'C\x01',
    'ć': 'c\x02', 'Ć': 'C\x02',
    'đ': 'd\x01', 'Đ': 'D\x01',
    'š': 's\x01', 'Š': 'S\x01',
    'ž': 'z\x01', 'Ž': 'Z\x01',
})


def _hr_key(text):
    """Croatian-aware sort key for a single string."""
    return text.lower().translate(_HR_MAP)


def _natural_key(text):
    """Natural sort key – numbers sort numerically, text alphabetically (Croatian)."""
    parts = re.split(r'(\d+)', text)
    return [(int(p) if p.isdigit() else _hr_key(p)) for p in parts]


def sort_classrooms(rows):
    """Sort classrooms naturally (numeric names first, then alphabetic)."""
    return sorted(rows, key=lambda r: _natural_key(r['name']))


def sort_professors(rows):
    """Sort professors by last_name, first_name (Croatian A-Z)."""
    return sorted(rows, key=lambda r: (_hr_key(r['last_name']), _hr_key(r['first_name'])))


def sort_programs(rows):
    """Sort study programs by name, element (Croatian A-Z)."""
    return sorted(rows, key=lambda r: (_hr_key(r['name']), _hr_key(r['element'] or '')))


def sort_courses(rows):
    """Sort courses by name (Croatian A-Z)."""
    return sorted(rows, key=lambda r: _hr_key(r['name']))


def read_excel_rows(file_storage):
    """Read rows from an uploaded Excel file (.xlsx or .xls).
    Returns a list of tuples (row values).
    """
    data = file_storage.read()
    filename = file_storage.filename or ''

    if filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheet_by_index(0)
        rows = []
        for i in range(ws.nrows):
            rows.append(tuple(ws.cell_value(i, j) for j in range(ws.ncols)))
        return rows
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows

DAYS = {
    1: 'Ponedjeljak',
    2: 'Utorak',
    3: 'Srijeda',
    4: 'Četvrtak',
    5: 'Petak',
    6: 'Subota',
    7: 'Nedjelja',
}

DAY_STATUSES = ['neradni', 'praznik', 'nenastavni']

TIME_SLOTS_REDOVITI = [
    '08:00 - 08:45', '08:45 - 09:30',
    '10:00 - 10:45', '10:45 - 11:30',
    '12:00 - 12:45', '12:45 - 13:30',
    '14:00 - 14:45', '14:45 - 15:30',
    '16:00 - 16:45', '16:45 - 17:30',
    '18:00 - 18:45', '18:45 - 19:30',
]

TIME_SLOTS_IZVANREDNI = [
    '08:30 - 09:15', '09:15 - 10:00', '10:00 - 10:45', '10:45 - 11:30',
    '11:30 - 12:15', '12:15 - 13:00', '13:00 - 13:45', '13:45 - 14:30',
    '14:30 - 15:15', '15:45 - 16:30', '16:30 - 17:15', '17:15 - 18:00',
    '18:00 - 18:45', '18:45 - 19:30', '19:30 - 20:15', '20:15 - 21:00',
]

# Backward compat alias
TIME_SLOTS = TIME_SLOTS_REDOVITI

TIMES_REDOVITI = [
    '08:00', '08:45', '09:30', '10:00', '10:45', '11:30',
    '12:00', '12:45', '13:30', '14:00', '14:45', '15:30',
    '16:00', '16:45', '17:30', '18:00', '18:45', '19:30',
]

TIMES_IZVANREDNI = [
    '08:30', '09:15', '10:00', '10:45', '11:30', '12:15',
    '13:00', '13:45', '14:30', '15:15', '15:45', '16:30',
    '17:15', '18:00', '18:45', '19:30', '20:15', '21:00',
]

# Backward compat alias
TIMES = TIMES_REDOVITI


def get_time_slots(study_mode=None):
    """Vrati vremenske slotove za prikaz ovisno o načinu studija."""
    if study_mode == 'izvanredni':
        return TIME_SLOTS_IZVANREDNI
    return TIME_SLOTS_REDOVITI


def get_times(study_mode=None):
    """Vrati listu mogućih vremena za start/end odabir."""
    if study_mode == 'izvanredni':
        return TIMES_IZVANREDNI
    return TIMES_REDOVITI

WEEK_TYPES = ['kontinuirano', '1. tjedan', '2. tjedan']

GROUPS = ['A', 'B', 'C', 'D', 'E']

MODULES = ['A', 'B', 'C']

SEMESTER_TYPES = ['zimski', 'ljetni']

TEACHING_FORMS = ['predavanja', 'seminari', 'vježbe']

STUDY_MODES = ['redoviti', 'izvanredni']

DAYS_REDOVITI = {1: 'Ponedjeljak', 2: 'Utorak', 3: 'Srijeda', 4: 'Četvrtak', 5: 'Petak'}
DAYS_IZVANREDNI = {4: 'Četvrtak', 5: 'Petak', 6: 'Subota'}


DAYS_ALL = {1: 'Ponedjeljak', 2: 'Utorak', 3: 'Srijeda', 4: 'Četvrtak', 5: 'Petak', 6: 'Subota'}


def get_display_days(study_mode):
    """Vrati odgovarajući DAYS dict prema načinu studija."""
    if study_mode == 'izvanredni':
        return DAYS_IZVANREDNI
    if study_mode == 'redoviti':
        return DAYS_REDOVITI
    return DAYS_ALL


def get_week_dates(ref_date_str, study_mode):
    """Iz referentnog datuma izračunaj datume za dane u tjednu.

    Za izvanredne: vraća {4: 'dd.mm.yyyy.', 5: '...', 6: '...'}
    za čet-pet-sub tjedna u kojem je ref_date.
    """
    d = datetime.strptime(ref_date_str, '%Y-%m-%d')
    iso_day = d.isoweekday()  # 1=pon, 7=ned
    monday = d - timedelta(days=iso_day - 1)

    display_days = get_display_days(study_mode)
    day_dates = {}
    for day_num in display_days:
        day_date = monday + timedelta(days=day_num - 1)
        day_dates[day_num] = day_date.strftime('%d.%m.%Y.')
    return day_dates


def get_week_date_range(ref_date_str, study_mode):
    """Iz referentnog datuma izračunaj date_from i date_to za filtriranje.

    Za izvanredne: vraća (YYYY-MM-DD četvrtak, YYYY-MM-DD subota) tog tjedna.
    """
    d = datetime.strptime(ref_date_str, '%Y-%m-%d')
    iso_day = d.isoweekday()
    monday = d - timedelta(days=iso_day - 1)

    display_days = get_display_days(study_mode)
    day_nums = sorted(display_days.keys())
    date_from = (monday + timedelta(days=day_nums[0] - 1)).strftime('%Y-%m-%d')
    date_to = (monday + timedelta(days=day_nums[-1] - 1)).strftime('%Y-%m-%d')
    return date_from, date_to

def group_entries_by_week(entries, study_mode='izvanredni'):
    """Grupiraj stavke po tjednima na temelju datuma.

    Vraća listu dicts sortiranu kronološki:
    [{'label': 'Tjedan: 06.02. - 08.02.2025.', 'day_dates': {...}, 'entries': [...]}]
    """
    from collections import defaultdict
    display_days = get_display_days(study_mode)
    day_nums = sorted(display_days.keys())

    # Grupiraj po ISO tjednu (year, week_number)
    week_groups = defaultdict(list)
    for e in entries:
        if not e['date']:
            continue
        d = datetime.strptime(e['date'], '%Y-%m-%d')
        iso_year, iso_week, _ = d.isocalendar()
        week_groups[(iso_year, iso_week)].append(e)

    result = []
    for key in sorted(week_groups.keys()):
        week_entries = week_groups[key]
        # Izračunaj ponedjeljak tog tjedna
        ref = datetime.strptime(week_entries[0]['date'], '%Y-%m-%d')
        monday = ref - timedelta(days=ref.isoweekday() - 1)

        day_dates = {}
        for day_num in display_days:
            day_date = monday + timedelta(days=day_num - 1)
            day_dates[day_num] = day_date.strftime('%d.%m.%Y.')

        # Label: "dd.mm. - dd.mm.yyyy."
        first_day = monday + timedelta(days=day_nums[0] - 1)
        last_day = monday + timedelta(days=day_nums[-1] - 1)
        label = f"{first_day.strftime('%d.%m.')} - {last_day.strftime('%d.%m.%Y.')}"

        result.append({
            'label': label,
            'day_dates': day_dates,
            'entries': week_entries,
        })

    return result


PROFESSOR_TITLES = [
    '', 'prof. dr. sc.', 'izv. prof. dr. sc.', 'doc. dr. sc.',
    'prof. dr. art.', 'izv. prof. dr. art.',
    'v. asis.', 'asis.', 'v. pred.', 'pred.',
]

def _hsl_to_hex(h, s, l):
    """Pretvori HSL (h:0-360, s:0-1, l:0-1) u hex string."""
    import colorsys
    r, g, b = colorsys.hls_to_rgb(h / 360.0, l, s)
    return '#{:02x}{:02x}{:02x}'.format(int(r * 255), int(g * 255), int(b * 255))


def _generate_colors(n=200):
    """Generiraj n razlicitih boja koristeci HSL."""
    colors = []
    sat_levels = [0.30, 0.50, 0.40, 0.60]
    light_bg = [0.88, 0.92, 0.85, 0.90]
    for i in range(n):
        hue = (i * 137.508) % 360  # zlatni kut za distribuciju
        variant = i % len(sat_levels)
        colors.append({
            'bg': _hsl_to_hex(hue, sat_levels[variant], light_bg[variant]),
            'border': _hsl_to_hex(hue, 0.70, 0.40),
            'text': _hsl_to_hex(hue, 0.60, 0.20),
        })
    return colors


COLORS_PALETTE = _generate_colors(200)


def build_program_colors(entries):
    """Izgradi mapiranje study_program_id -> boja iz palete.

    Koristi study_program_id kao indeks za konzistentne boje
    neovisno o kontekstu prikaza.
    """
    seen = set()
    for entry in entries:
        seen.add(entry['study_program_id'])
    return {pid: COLORS_PALETTE[(pid - 1) % len(COLORS_PALETTE)]
            for pid in seen}


def date_to_day_of_week(date_str):
    """Pretvori datum (YYYY-MM-DD) u dan u tjednu (1=pon, 7=ned)."""
    d = datetime.strptime(date_str, '%Y-%m-%d')
    # isoweekday(): Monday=1, Sunday=7
    return d.isoweekday()


def times_overlap(start_a, end_a, start_b, end_b):
    """Provjeri preklapaju li se dva vremenska raspona."""
    return start_a < end_b and end_a > start_b


def weeks_overlap(week_type_a, week_type_b):
    """Provjeri da li se dva tipa tjedna preklapaju."""
    if week_type_a == 'kontinuirano' or week_type_b == 'kontinuirano':
        return True
    return week_type_a == week_type_b


def check_conflicts(entry_data, exclude_id=None):
    """Provjeri konflikte za novi/editirani unos rasporeda."""
    db = get_db()
    conflicts = []

    # Provjeri study_mode za entry — izvanredni koristi datume, ne dane
    study_mode = None
    if entry_data.get('study_program_id'):
        sp = db.execute('SELECT study_mode FROM study_program WHERE id = ?',
                        (entry_data['study_program_id'],)).fetchone()
        if sp:
            study_mode = sp['study_mode']

    query = '''
        SELECT se.*, c.name as course_name, p.first_name, p.last_name, p.title,
               cl.name as classroom_name, sp.name as program_name, sp.element as program_element, sp.study_mode
        FROM schedule_entry se
        JOIN course c ON se.course_id = c.id
        JOIN professor p ON se.professor_id = p.id
        JOIN classroom cl ON se.classroom_id = cl.id
        JOIN study_program sp ON se.study_program_id = sp.id
        WHERE se.academic_year_id = ?
        AND se.day_of_week = ?
        AND se.start_time < ?
        AND se.end_time > ?
    '''
    params = [
        entry_data['academic_year_id'],
        entry_data['day_of_week'],
        entry_data['end_time'],
        entry_data['start_time'],
    ]

    if exclude_id:
        query += ' AND se.id != ?'
        params.append(exclude_id)

    overlapping = db.execute(query, params).fetchall()

    # Filtriraj po datumu: dva izvanredna unosa na različite datume ne konfliktiiraju.
    # Redoviti unos (bez datuma) konfliktiira sa svima na isti day_of_week jer se ponavlja svaki tjedan.
    entry_date = entry_data.get('date')
    if entry_date:
        overlapping = [e for e in overlapping
                       if not e['date'] or e['date'] == entry_date]

    for existing in overlapping:
        if not weeks_overlap(entry_data['week_type'], existing['week_type']):
            continue

        week_info = f" ({existing['week_type']})" if existing['week_type'] != 'kontinuirano' else ''
        time_info = f"{existing['start_time']}-{existing['end_time']}"
        prog_label = f"{existing['program_name']} ({existing['study_mode']})"
        course_prog = f"{existing['course_name']} [{prog_label}]"

        if existing['professor_id'] == int(entry_data['professor_id']):
            prof_name = f"{existing['title']} {existing['first_name']} {existing['last_name']}".strip()
            conflicts.append(
                f"Profesor {prof_name} je već zauzet/a: "
                f"{course_prog} u {existing['classroom_name']} ({time_info}){week_info}"
            )

        if existing['classroom_id'] == int(entry_data['classroom_id']):
            conflicts.append(
                f"Učionica {existing['classroom_name']} je već zauzeta: "
                f"{course_prog} ({time_info}){week_info}"
            )

        if (entry_data.get('group_name')
                and existing['group_name'] == entry_data['group_name']
                and existing['study_program_id'] == int(entry_data['study_program_id'])
                and existing['semester_number'] == int(entry_data['semester_number'])):
            conflicts.append(
                f"Grupa {entry_data['group_name']} ({prog_label}, "
                f"{existing['semester_number']}. semestar) već ima predavanje: "
                f"{course_prog} ({time_info}){week_info}"
            )

    return conflicts


def build_timetable_grid(entries, days=None, time_slots=None):
    """Izgradi grid strukturu za prikaz rasporeda.

    Stavka se prikazuje u svakom time slotu koji pokriva.
    days: dict {day_num: day_name} - koji dani se prikazuju (default: svi 1-7)
    time_slots: lista vremenskih slotova (default: TIME_SLOTS_REDOVITI)
    Returns dict: {time_slot: {day_of_week: [entries]}}
    """
    if days is None:
        days = DAYS
    if time_slots is None:
        time_slots = TIME_SLOTS_REDOVITI
    grid = {}
    for ts in time_slots:
        grid[ts] = {}
        for day in days:
            grid[ts][day] = []

    for entry in entries:
        day = entry['day_of_week']
        if day not in days:
            continue
        e_start = entry['start_time']
        e_end = entry['end_time']

        placed = False
        for ts in time_slots:
            slot_start, slot_end = ts.split(' - ')
            if e_start < slot_end and e_end > slot_start:
                grid[ts][day].append(entry)
                placed = True

        if not placed:
            # Stavka pada u pauzu između slotova (npr. 11:30-12:00)
            # Postavi je u prvi slot koji počinje na ili nakon završetka stavke
            for ts in time_slots:
                slot_start = ts.split(' - ')[0]
                if slot_start >= e_end:
                    grid[ts][day].append(entry)
                    placed = True
                    break
            if not placed:
                # Stavka je nakon svih slotova — stavi u zadnji slot
                grid[time_slots[-1]][day].append(entry)

    return grid


def build_day_dates(entries, display_days=None):
    """Izgradi mapiranje dan -> datum string iz stavki rasporeda.

    Vraća {day_num: 'dd.mm.yyyy.'} ili {day_num: 'dd.mm., dd.mm.'} ako
    ima više tjedana. Koristi samo dane iz display_days ako je zadano.
    """
    day_dates = {}
    unique_dates = set()
    for entry in entries:
        if entry['date']:
            unique_dates.add(entry['date'])

    if not unique_dates:
        return {}

    for date_str in unique_dates:
        d = datetime.strptime(date_str, '%Y-%m-%d')
        iso_day = d.isoweekday()
        monday = d - timedelta(days=iso_day - 1)
        days_to_show = display_days or DAYS
        for day_num in days_to_show:
            day_date = monday + timedelta(days=day_num - 1)
            if day_num not in day_dates:
                day_dates[day_num] = set()
            day_dates[day_num].add(day_date.strftime('%d.%m.%Y.'))

    return {day: ', '.join(sorted(dates)) for day, dates in day_dates.items()}


def compute_day_columns(entries, days):
    """Izračunaj broj pod-stupaca po danu i dodijeli svaki unos tracku.

    Koristi greedy algoritam za track assignment:
    unosi na istom tracku se ne preklapaju vremenski.

    Za dane s mješavinom '1. tjedan' / '2. tjedan' unosa, koristi
    week-aware raspored: track 0 = 1. tjedan, track 1 = 2. tjedan.

    Returns:
        day_columns: {day_num: int} - broj pod-stupaca za svaki dan (min 1)
        entry_tracks: {entry_id: int} - indeks tracka za svaki unos
        week_split_days: set - dani koji su splitani po tjednima
    """
    day_entries = {day: [] for day in days}
    seen = set()
    for entry in entries:
        if entry['id'] not in seen and entry['day_of_week'] in days:
            seen.add(entry['id'])
            day_entries[entry['day_of_week']].append(entry)

    day_columns = {}
    entry_tracks = {}
    week_split_days = set()

    for day in days:
        ents = sorted(day_entries[day], key=lambda e: (e['start_time'], e['end_time']))
        if not ents:
            day_columns[day] = 1
            continue

        # Provjeri ima li dan unose s specifičnim tjednima
        week_types = set(e['week_type'] for e in ents)
        has_specific_weeks = bool(week_types - {'kontinuirano'})

        if has_specific_weeks:
            # Week-split: 2 pod-stupca (1. tjedan | 2. tjedan)
            week_split_days.add(day)
            for entry in ents:
                if entry['week_type'] == '2. tjedan':
                    entry_tracks[entry['id']] = 1
                else:  # '1. tjedan' ili 'kontinuirano'
                    entry_tracks[entry['id']] = 0
            day_columns[day] = 2
        else:
            # Grupiraj stavke s identičnim (start_time, end_time) u jednu "scheduling unit"
            from collections import defaultdict
            time_groups = defaultdict(list)
            for entry in ents:
                key = (entry['start_time'], entry['end_time'])
                time_groups[key].append(entry)

            units = sorted(time_groups.items(), key=lambda x: x[0])

            # Greedy track assignment na unitima (ne na pojedinačnim stavkama)
            track_ends = []
            for (st, et), group_entries in units:
                placed = False
                for i in range(len(track_ends)):
                    if track_ends[i] <= st:
                        track_ends[i] = et
                        for entry in group_entries:
                            entry_tracks[entry['id']] = i
                        placed = True
                        break
                if not placed:
                    track_idx = len(track_ends)
                    for entry in group_entries:
                        entry_tracks[entry['id']] = track_idx
                    track_ends.append(et)

            day_columns[day] = max(len(track_ends), 1)

    # Kad postoje week-split dani, osiguraj da SVI dani imaju 2 stupca
    # (inače table-layout:fixed daje splitanim danima dvostruku širinu)
    if week_split_days:
        for day in days:
            if day_columns.get(day, 1) < 2:
                day_columns[day] = 2

    return day_columns, entry_tracks, week_split_days


def build_cell_info(grid, time_slots, days, day_columns=None, entry_tracks=None, week_split_days=None):
    """Izgradi info za renderiranje celija s track-based pod-stupcima.

    Returns dict: {time_slot: {day: [{'skip': bool, 'rowspan': int, 'colspan': int, 'entries': list}]}}
    Svaki dan ima listu od day_columns[day] elemenata (trackova).
    Za week-split dane, 'kontinuirano' unosi dobivaju colspan=2.
    """
    ts_list = list(time_slots)
    slot_bounds = [(ts.split(' - ')[0], ts.split(' - ')[1]) for ts in ts_list]

    if day_columns is None:
        day_columns = {day: 1 for day in days}
    if entry_tracks is None:
        entry_tracks = {}
    if week_split_days is None:
        week_split_days = set()

    # Dani koji su padani na 2 stupca ali NISU pravi week-split
    # (sve njihove ćelije dobivaju colspan=2, track 1 je skip)
    padded_days = set()
    if week_split_days:
        for day in days:
            if day not in week_split_days and day_columns.get(day, 1) > 1:
                padded_days.add(day)

    cell_info = {}
    for ts in ts_list:
        cell_info[ts] = {}
        for day in days:
            n = day_columns.get(day, 1)
            if day in padded_days:
                # Padani dan: track 0 ima colspan=n, trackovi 1+ su skip
                cell_info[ts][day] = [
                    {'skip': False, 'rowspan': 1, 'colspan': n, 'entries': []}
                ] + [
                    {'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': []}
                    for _ in range(1, n)
                ]
            else:
                cell_info[ts][day] = [
                    {'skip': False, 'rowspan': 1, 'colspan': 1, 'entries': []}
                    for _ in range(n)
                ]

    for day in days:
        seen = set()
        day_ents = []
        for ts in ts_list:
            for entry in grid[ts][day]:
                if entry['id'] not in seen:
                    seen.add(entry['id'])
                    day_ents.append(entry)

        for entry in day_ents:
            track = entry_tracks.get(entry['id'], 0)

            # Na padanim danima (nisu pravi week-split, ali imaju 2 stupca)
            # sve stavke moraju ići na track 0 jer track 1 je skip
            if day in padded_days and track > 0:
                track = 0

            # Colspan: kontinuirano na week-split danima, ili sve na padanim danima
            colspan = 1
            if day in week_split_days and entry['week_type'] == 'kontinuirano':
                # Samo colspan=2 ako NEMA preklapajućih unosa na drugim trackovima
                has_overlap = False
                for other in day_ents:
                    if other['id'] != entry['id']:
                        other_track = entry_tracks.get(other['id'], 0)
                        if other_track != track:
                            if (other['start_time'] < entry['end_time'] and
                                    other['end_time'] > entry['start_time']):
                                has_overlap = True
                                break
                if not has_overlap:
                    colspan = day_columns.get(day, 1)
            elif day in padded_days:
                colspan = day_columns.get(day, 1)

            start_idx = None
            span = 0
            for idx, (s, e) in enumerate(slot_bounds):
                if entry['start_time'] < e and entry['end_time'] > s:
                    if start_idx is None:
                        start_idx = idx
                    span += 1

            if start_idx is None:
                # Stavka pada u pauzu između slotova — pronađi najbliži slot
                for idx, (s, e) in enumerate(slot_bounds):
                    if s >= entry['end_time']:
                        start_idx = idx
                        span = 1
                        break
                if start_idx is None and slot_bounds:
                    start_idx = len(slot_bounds) - 1
                    span = 1

            if start_idx is not None:
                ts = ts_list[start_idx]
                existing = cell_info[ts][day][track]

                # Ćelija je skip (pokrivena rowspan-om drugog unosa) —
                # pronađi roditeljsku ćeliju i mergaj stavku tamo
                if existing.get('skip', False):
                    parent_found = False
                    for back_idx in range(start_idx - 1, -1, -1):
                        parent = cell_info[ts_list[back_idx]][day][track]
                        if not parent.get('skip', False) and parent.get('entries'):
                            parent['entries'].append(entry)
                            # Proširi rowspan da pokrije i novu stavku
                            needed_span = start_idx - back_idx + span
                            if needed_span > parent['rowspan']:
                                old_parent_span = parent['rowspan']
                                parent['rowspan'] = needed_span
                                for k in range(back_idx + old_parent_span, back_idx + needed_span):
                                    if k < len(ts_list):
                                        cell_info[ts_list[k]][day][track] = {
                                            'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                        }
                            parent_found = True
                            break
                    if parent_found:
                        continue
                    # Nema roditelja — ne gubi stavku, stavi je u ovu ćeliju
                    cell_info[ts][day][track] = {
                        'skip': False, 'rowspan': span, 'colspan': colspan, 'entries': [entry],
                    }
                    for k in range(start_idx + 1, start_idx + span):
                        if k < len(ts_list):
                            target = cell_info[ts_list[k]][day][track]
                            if not target.get('entries'):
                                cell_info[ts_list[k]][day][track] = {
                                    'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                }
                    continue

                if existing['entries']:
                    # Isti track, preklapajuće vrijeme → UVIJEK merge
                    # Koristi max rowspan i max colspan (ne zamjenjuj!)
                    existing['entries'].append(entry)
                    old_span = existing['rowspan']
                    old_colspan = existing.get('colspan', 1)
                    new_span = max(old_span, span)
                    new_colspan = max(old_colspan, colspan)
                    existing['rowspan'] = new_span
                    existing['colspan'] = new_colspan

                    # Proširi rowspan skip markere ako je span porastao
                    if new_span > old_span:
                        for k in range(start_idx + old_span, start_idx + new_span):
                            if k < len(ts_list):
                                target = cell_info[ts_list[k]][day][track]
                                if not target.get('entries'):
                                    cell_info[ts_list[k]][day][track] = {
                                        'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                    }

                    # Proširi colspan skip markere ako je colspan porastao
                    if new_colspan > old_colspan:
                        for t in range(track + old_colspan, track + new_colspan):
                            if t < day_columns.get(day, 1):
                                for k in range(start_idx, start_idx + new_span):
                                    if k < len(ts_list):
                                        target_c = cell_info[ts_list[k]][day][t]
                                        if not target_c.get('entries'):
                                            cell_info[ts_list[k]][day][t] = {
                                                'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                            }

                    # Ako je span porastao i stari colspan > 1, proširi colspan skip na nove redove
                    if new_span > old_span and old_colspan > 1:
                        for t in range(track + 1, track + old_colspan):
                            if t < day_columns.get(day, 1):
                                for k in range(start_idx + old_span, start_idx + new_span):
                                    if k < len(ts_list):
                                        target_c = cell_info[ts_list[k]][day][t]
                                        if not target_c.get('entries'):
                                            cell_info[ts_list[k]][day][t] = {
                                                'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                            }
                else:
                    cell_info[ts][day][track] = {
                        'skip': False,
                        'rowspan': span,
                        'colspan': colspan,
                        'entries': [entry],
                    }
                    # Skip ćelije za rowspan
                    for k in range(start_idx + 1, start_idx + span):
                        if k < len(ts_list):
                            target = cell_info[ts_list[k]][day][track]
                            if not target.get('entries'):
                                cell_info[ts_list[k]][day][track] = {
                                    'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                }
                    # Skip ćelije za colspan
                    if colspan > 1:
                        for t in range(track + 1, track + colspan):
                            if t < day_columns.get(day, 1):
                                for k in range(start_idx, start_idx + span):
                                    if k < len(ts_list):
                                        target = cell_info[ts_list[k]][day][t]
                                        if not target.get('entries'):
                                            cell_info[ts_list[k]][day][t] = {
                                                'skip': True, 'rowspan': 1, 'colspan': 1, 'entries': [],
                                            }

    # Sigurnosna provjera: osiguraj da NIJEDNA stavka nije izgubljena
    for day in days:
        grid_ids = set()
        for ts in ts_list:
            for entry in grid[ts][day]:
                grid_ids.add(entry['id'])
        ci_ids = set()
        for ts in ts_list:
            for track_info in cell_info[ts][day]:
                for entry in track_info['entries']:
                    ci_ids.add(entry['id'])
        missing = grid_ids - ci_ids
        if missing:
            # Pronađi izgubljene stavke i dodaj ih u prvu dostupnu ćeliju
            missing_entries = {}
            for ts in ts_list:
                for entry in grid[ts][day]:
                    if entry['id'] in missing and entry['id'] not in missing_entries:
                        missing_entries[entry['id']] = entry
            for entry in missing_entries.values():
                e_start = entry['start_time']
                placed = False
                for ts_idx, ts in enumerate(ts_list):
                    s, e = slot_bounds[ts_idx]
                    if e_start < e and entry['end_time'] > s:
                        for t in range(day_columns.get(day, 1)):
                            ci_cell = cell_info[ts][day][t]
                            if not ci_cell.get('skip', False):
                                ci_cell['entries'].append(entry)
                                placed = True
                                break
                        if placed:
                            break
                if not placed and ts_list:
                    # Zadnje sredstvo: dodaj u prvu ćeliju dana
                    cell_info[ts_list[0]][day][0]['entries'].append(entry)

    return cell_info


def get_schedule_entries(filters):
    """Dohvati stavke rasporeda s filterima."""
    db = get_db()

    query = '''
        SELECT se.*, c.name as course_name, c.code as course_code,
               p.first_name, p.last_name, p.title,
               cl.name as classroom_name,
               sp.name as program_name, sp.code as program_code, sp.element as program_element,
               sp.study_mode,
               ay.name as academic_year_name
        FROM schedule_entry se
        JOIN course c ON se.course_id = c.id
        JOIN professor p ON se.professor_id = p.id
        JOIN classroom cl ON se.classroom_id = cl.id
        JOIN study_program sp ON se.study_program_id = sp.id
        JOIN academic_year ay ON se.academic_year_id = ay.id
        WHERE 1=1
    '''
    params = []

    if filters.get('academic_year_id'):
        query += ' AND se.academic_year_id = ?'
        params.append(filters['academic_year_id'])
    if filters.get('study_program_id'):
        query += ' AND se.study_program_id = ?'
        params.append(filters['study_program_id'])
    if filters.get('semester_type'):
        query += ' AND se.semester_type = ?'
        params.append(filters['semester_type'])
    if filters.get('semester_number'):
        query += ' AND se.semester_number = ?'
        params.append(filters['semester_number'])
    if filters.get('professor_id'):
        query += ' AND se.professor_id = ?'
        params.append(filters['professor_id'])
    if filters.get('classroom_id'):
        query += ' AND se.classroom_id = ?'
        params.append(filters['classroom_id'])
    if filters.get('week_type'):
        if filters['week_type'] == 'kontinuirano':
            query += " AND se.week_type = 'kontinuirano'"
        elif filters['week_type'] == '1. tjedan':
            query += " AND se.week_type IN ('kontinuirano', '1. tjedan')"
        elif filters['week_type'] == '2. tjedan':
            query += " AND se.week_type IN ('kontinuirano', '2. tjedan')"
    if filters.get('study_mode'):
        query += ' AND sp.study_mode = ?'
        params.append(filters['study_mode'])
    if filters.get('date_from') and filters.get('date_to'):
        query += ' AND se.date >= ? AND se.date <= ?'
        params.append(filters['date_from'])
        params.append(filters['date_to'])
    if filters.get('published_only'):
        query += ' AND se.is_published = 1'

    query += ' ORDER BY se.day_of_week, se.start_time'
    return db.execute(query, params).fetchall()
