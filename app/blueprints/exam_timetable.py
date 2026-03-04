import io
from collections import defaultdict
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash
from app.auth import login_required, is_admin as check_admin
from app.db import get_db
from app.audit import log_audit
from app.models import DAYS_ALL, sort_classrooms, sort_professors
from app.blueprints.timetable import get_day_statuses

bp = Blueprint('exam_timetable', __name__)


def _format_date(iso_date):
    try:
        return datetime.strptime(iso_date, '%Y-%m-%d').strftime('%d.%m.%Y.')
    except (ValueError, TypeError):
        return iso_date or ''


def _get_default_academic_year_id():
    db = get_db()
    row = db.execute('SELECT id FROM academic_year WHERE is_default = 1').fetchone()
    return row['id'] if row else None


def _get_exam_entries(academic_year_id, published_only=False, professor_id=None, classroom_id=None):
    """Dohvati ispitne rokove s JOIN-ovima."""
    db = get_db()
    conditions = ['ee.academic_year_id = ?']
    params = [academic_year_id]

    if published_only:
        conditions.append('ee.is_published = 1')
    if professor_id:
        conditions.append('ee.professor_id = ?')
        params.append(professor_id)
    if classroom_id:
        conditions.append('ee.classroom_id = ?')
        params.append(classroom_id)

    where = ' AND '.join(conditions)
    return db.execute(f'''
        SELECT ee.*,
               p.first_name, p.last_name, p.title,
               cl.name as classroom_name,
               ay.name as academic_year_name
        FROM exam_entry ee
        JOIN professor p ON ee.professor_id = p.id
        JOIN classroom cl ON ee.classroom_id = cl.id
        JOIN academic_year ay ON ee.academic_year_id = ay.id
        WHERE {where}
        ORDER BY ee.date, ee.start_time
    ''', params).fetchall()


def _group_by_week(entries):
    """Grupiraj ispitne rokove po ISO tjednima s day_dates za prikaz praznika."""
    week_groups = defaultdict(list)
    for e in entries:
        if not e['date']:
            continue
        d = datetime.strptime(e['date'], '%Y-%m-%d')
        iso_year, iso_week, _ = d.isocalendar()
        week_groups[(iso_year, iso_week)].append(e)

    result = []
    for key in sorted(week_groups.keys()):
        week_entries = week_groups[key]
        ref = datetime.strptime(week_entries[0]['date'], '%Y-%m-%d')
        monday = ref - timedelta(days=ref.isoweekday() - 1)
        saturday = monday + timedelta(days=5)
        label = f"{monday.strftime('%d.%m.')} - {saturday.strftime('%d.%m.%Y.')}"

        # day_dates: {1: '02.03.2026.', 2: '03.03.2026.', ..., 6: '07.03.2026.'}
        day_dates = {}
        for day_num in DAYS_ALL:
            day_date = monday + timedelta(days=day_num - 1)
            day_dates[day_num] = day_date.strftime('%d.%m.%Y.')

        # Grupiraj po danima unutar tjedna
        by_day = defaultdict(list)
        for e in week_entries:
            by_day[e['day_of_week']].append(e)

        result.append({
            'label': label,
            'entries': week_entries,
            'by_day': dict(by_day),
            'day_dates': day_dates,
        })
    return result


def _parse_schedule_date(raw):
    """Parsiraj datum iz filtera (dd.mm.YYYY. ili YYYY-MM-DD) u ISO format."""
    if not raw:
        return None, None
    if '.' in raw:
        parts = raw.replace('.', ' ').split()
        if len(parts) >= 3:
            try:
                d = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                return d.strftime('%Y-%m-%d'), raw
            except (ValueError, IndexError):
                return None, None
    else:
        try:
            d = datetime.strptime(raw, '%Y-%m-%d')
            return raw, d.strftime('%d.%m.%Y.')
        except ValueError:
            return None, None
    return None, None


@bp.route('/')
def index():
    db = get_db()
    academic_year_id = request.args.get('academic_year_id', type=int)
    if not academic_year_id:
        academic_year_id = _get_default_academic_year_id()

    professor_id = request.args.get('professor_id', type=int)
    classroom_id = request.args.get('classroom_id', type=int)

    # Datum filter — prikaži samo tjedan koji sadrži odabrani datum
    schedule_date_raw = request.args.get('schedule_date', '').strip()
    schedule_date_iso, schedule_date_display = _parse_schedule_date(schedule_date_raw)

    is_admin = check_admin()
    published_only = not is_admin

    entries = _get_exam_entries(academic_year_id, published_only=published_only,
                                professor_id=professor_id, classroom_id=classroom_id) if academic_year_id else []
    weeks = _group_by_week(entries)

    # Filtriraj na jedan tjedan ako je odabran datum
    if schedule_date_iso and weeks:
        target = datetime.strptime(schedule_date_iso, '%Y-%m-%d')
        target_year, target_week, _ = target.isocalendar()
        filtered = []
        for week in weeks:
            ref = datetime.strptime(week['entries'][0]['date'], '%Y-%m-%d')
            wy, ww, _ = ref.isocalendar()
            if wy == target_year and ww == target_week:
                filtered.append(week)
                break
        weeks = filtered

    # Dohvati day_statuses (praznici) za svaki tjedan
    for week in weeks:
        week['day_statuses'] = get_day_statuses(academic_year_id, week['day_dates']) if academic_year_id else {}

    academic_years = db.execute('SELECT * FROM academic_year ORDER BY name DESC').fetchall()
    professors = sort_professors(db.execute('SELECT * FROM professor').fetchall())
    classrooms = sort_classrooms(db.execute('SELECT * FROM classroom').fetchall())

    return render_template('exam_timetable/view.html',
                           weeks=weeks,
                           days=DAYS_ALL,
                           academic_years=academic_years,
                           professors=professors,
                           classrooms=classrooms,
                           selected_year=academic_year_id,
                           selected_professor=professor_id,
                           selected_classroom=classroom_id,
                           selected_date=schedule_date_display or '',
                           is_admin=is_admin)


@bp.route('/excel')
def excel():
    academic_year_id = request.args.get('academic_year_id', type=int)
    if not academic_year_id:
        academic_year_id = _get_default_academic_year_id()
    if not academic_year_id:
        flash('Odaberite akademsku godinu.', 'warning')
        return redirect(url_for('exam_timetable.index'))

    professor_id = request.args.get('professor_id', type=int)
    classroom_id = request.args.get('classroom_id', type=int)

    is_admin = check_admin()
    published_only = not is_admin

    entries = _get_exam_entries(academic_year_id, published_only=published_only,
                                professor_id=professor_id, classroom_id=classroom_id)
    weeks = _group_by_week(entries)

    # Datum filter za Excel
    schedule_date_raw = request.args.get('schedule_date', '').strip()
    schedule_date_iso, _ = _parse_schedule_date(schedule_date_raw)
    if schedule_date_iso and weeks:
        target = datetime.strptime(schedule_date_iso, '%Y-%m-%d')
        target_year, target_week, _ = target.isocalendar()
        weeks = [w for w in weeks
                 if (lambda r: r.isocalendar()[:2] == (target_year, target_week))(
                     datetime.strptime(w['entries'][0]['date'], '%Y-%m-%d'))]

    db = get_db()
    ay = db.execute('SELECT name FROM academic_year WHERE id = ?', (academic_year_id,)).fetchone()
    ay_name = ay['name'] if ay else ''

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    title_font = Font(bold=True, size=14)
    day_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    day_font = Font(bold=True, size=11)
    conflict_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    if not weeks:
        ws = wb.active
        ws.title = 'Ispitni rokovi'
        ws.cell(row=1, column=1, value='Nema ispitnih rokova za odabranu akademsku godinu.').font = title_font
    else:
        first = True
        for week in weeks:
            sheet_name = week['label'][:31]
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            # Title
            title = f"Ispitni rokovi — {ay_name} — {week['label']}"
            ws.cell(row=1, column=1, value=title).font = title_font
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

            # Headers
            headers = ['Datum', 'Vrijeme', 'Tip', 'Profesor', 'Učionica', 'Napomena']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row = 4
            sorted_entries = sorted(week['entries'], key=lambda e: (e['date'], e['start_time']))
            current_day = None
            for e in sorted_entries:
                if e['day_of_week'] != current_day:
                    current_day = e['day_of_week']
                    day_name = DAYS_ALL.get(e['day_of_week'], '')
                    day_date = _format_date(e['date'])
                    day_cell = ws.cell(row=row, column=1, value=f"{day_name}, {day_date}")
                    day_cell.font = day_font
                    day_cell.fill = day_fill
                    for c in range(1, 7):
                        ws.cell(row=row, column=c).fill = day_fill
                        ws.cell(row=row, column=c).border = thin_border
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    row += 1

                prof = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
                values = [
                    _format_date(e['date']),
                    f"{e['start_time']}-{e['end_time']}",
                    e['exam_type'],
                    prof,
                    e['classroom_name'],
                    e['note'] or '',
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    if e['has_conflict']:
                        cell.fill = conflict_fill
                row += 1

            # Column widths
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 26
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ispitni_rokovi_{ay_name.replace('/', '-')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/unpublished')
@login_required
def unpublished():
    db = get_db()
    academic_year_id = request.args.get('academic_year_id', type=int)
    if not academic_year_id:
        academic_year_id = _get_default_academic_year_id()

    entries = db.execute('''
        SELECT ee.*,
               p.first_name, p.last_name, p.title,
               cl.name as classroom_name
        FROM exam_entry ee
        JOIN professor p ON ee.professor_id = p.id
        JOIN classroom cl ON ee.classroom_id = cl.id
        WHERE ee.is_published = 0 AND ee.academic_year_id = ?
        ORDER BY ee.date, ee.start_time
    ''', (academic_year_id,)).fetchall()

    entry_list = []
    for e in entries:
        prof = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
        entry_list.append({
            'id': e['id'],
            'date': _format_date(e['date']),
            'day_name': DAYS_ALL.get(e['day_of_week'], ''),
            'exam_type': e['exam_type'],
            'start_time': e['start_time'],
            'end_time': e['end_time'],
            'professor': prof,
            'classroom': e['classroom_name'],
            'note': e['note'] or '',
            'has_conflict': e['has_conflict'],
        })

    academic_years = db.execute('SELECT * FROM academic_year ORDER BY name DESC').fetchall()

    return render_template('exam_timetable/unpublished.html',
                           entries=entry_list,
                           academic_years=academic_years,
                           selected_year=academic_year_id)


@bp.route('/publish-selected', methods=['POST'])
@login_required
def publish_selected():
    db = get_db()
    if 'publish_all' in request.form:
        count = db.execute('SELECT COUNT(*) FROM exam_entry WHERE is_published = 0').fetchone()[0]
        if count > 0:
            db.execute('UPDATE exam_entry SET is_published = 1 WHERE is_published = 0')
            log_audit('publish', 'exam_entry', f'Objavljeno {count} ispitnih rokova')
            db.commit()
            flash(f'Uspješno objavljeno {count} ispitnih rokova.', 'success')
        else:
            flash('Nema neobjavljenih ispitnih rokova.', 'info')
    else:
        entry_ids = request.form.getlist('entry_ids', type=int)
        if entry_ids:
            placeholders = ','.join('?' * len(entry_ids))
            count = db.execute(
                f'UPDATE exam_entry SET is_published = 1 WHERE id IN ({placeholders}) AND is_published = 0',
                entry_ids
            ).rowcount
            if count > 0:
                log_audit('publish', 'exam_entry', f'Objavljeno {count} odabranih ispitnih rokova')
                db.commit()
                flash(f'Uspješno objavljeno {count} ispitnih rokova.', 'success')
            else:
                db.commit()
                flash('Odabrani ispitni rokovi su već objavljeni.', 'info')
        else:
            flash('Niste odabrali nijedan ispitni rok.', 'warning')
    return redirect(url_for('exam_timetable.unpublished'))
