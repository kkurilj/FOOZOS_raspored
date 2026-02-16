from flask import Blueprint, render_template, request, redirect, url_for, flash
from app.db import get_db
from app.models import PROFESSOR_TITLES

bp = Blueprint('professor', __name__)


@bp.route('/')
def index():
    db = get_db()
    professors = db.execute(
        'SELECT * FROM professor ORDER BY last_name, first_name'
    ).fetchall()
    return render_template('professor/index.html', professors=professors)


@bp.route('/create', methods=['GET', 'POST'])
def create():
    if request.method == 'POST':
        first_name = request.form['first_name'].strip()
        last_name = request.form['last_name'].strip()
        title = request.form['title'].strip()
        if not first_name or not last_name:
            flash('Ime i prezime su obavezni.', 'danger')
        else:
            db = get_db()
            db.execute(
                'INSERT INTO professor (first_name, last_name, title) VALUES (?, ?, ?)',
                (first_name, last_name, title)
            )
            db.commit()
            flash(f'Profesor "{title} {first_name} {last_name}" je dodan.'.strip(), 'success')
            return redirect(url_for('professor.index'))
    return render_template('professor/form.html', titles=PROFESSOR_TITLES)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
def edit(id):
    db = get_db()
    professor = db.execute('SELECT * FROM professor WHERE id = ?', (id,)).fetchone()
    if professor is None:
        flash('Profesor nije pronađen.', 'danger')
        return redirect(url_for('professor.index'))

    if request.method == 'POST':
        first_name = request.form['first_name'].strip()
        last_name = request.form['last_name'].strip()
        title = request.form['title'].strip()
        if not first_name or not last_name:
            flash('Ime i prezime su obavezni.', 'danger')
        else:
            db.execute(
                'UPDATE professor SET first_name = ?, last_name = ?, title = ? WHERE id = ?',
                (first_name, last_name, title, id)
            )
            db.commit()
            flash('Profesor je ažuriran.', 'success')
            return redirect(url_for('professor.index'))
    return render_template('professor/form.html', professor=professor, titles=PROFESSOR_TITLES)


@bp.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    db = get_db()
    db.execute('DELETE FROM professor WHERE id = ?', (id,))
    db.commit()
    flash('Profesor je obrisan.', 'success')
    return redirect(url_for('professor.index'))


@bp.route('/import', methods=['POST'])
def import_bulk():
    from openpyxl import load_workbook
    import io

    if 'file' not in request.files:
        flash('Datoteka nije odabrana.', 'danger')
        return redirect(url_for('professor.index'))

    file = request.files['file']
    if file.filename == '':
        flash('Datoteka nije odabrana.', 'danger')
        return redirect(url_for('professor.index'))

    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True)
        ws = wb.active
        db = get_db()
        added = 0
        skipped = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            # Format: titula, ime, prezime
            cells = [str(c).strip() if c is not None else '' for c in row[:3]]
            if len(cells) == 2:
                title, first_name, last_name = '', cells[0], cells[1]
            else:
                title, first_name, last_name = cells[0], cells[1], cells[2]

            if not first_name or not last_name:
                skipped += 1
                continue

            # Skip header row
            if first_name.lower() in ('ime', 'first_name', 'name') and last_name.lower() in ('prezime', 'last_name', 'surname'):
                continue

            db.execute(
                'INSERT INTO professor (first_name, last_name, title) VALUES (?, ?, ?)',
                (first_name, last_name, title)
            )
            added += 1

        db.commit()
        wb.close()
        flash(f'Uvezeno {added} profesora.' + (f' Preskočeno {skipped} redaka.' if skipped else ''), 'success')
    except Exception as e:
        flash(f'Greška pri uvozu: {e}', 'danger')

    return redirect(url_for('professor.index'))
