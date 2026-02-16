from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from app.db import get_db

bp = Blueprint('course', __name__)


@bp.route('/')
def index():
    db = get_db()
    courses = db.execute('''
        SELECT c.*, sp.name as program_name, sp.code as program_code, sp.element as program_element
        FROM course c
        LEFT JOIN study_program sp ON c.study_program_id = sp.id
        ORDER BY sp.name, sp.element, c.name
    ''').fetchall()
    programs = db.execute('SELECT * FROM study_program ORDER BY name, element').fetchall()
    return render_template('course/index.html', courses=courses, programs=programs)


@bp.route('/create', methods=['GET', 'POST'])
def create():
    db = get_db()
    programs = db.execute('SELECT * FROM study_program ORDER BY name, element').fetchall()

    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip()
        study_program_id = request.form.get('study_program_id', type=int)
        if not name or not code or not study_program_id:
            flash('Sva polja su obavezna.', 'danger')
        else:
            try:
                db.execute(
                    'INSERT INTO course (name, code, study_program_id) VALUES (?, ?, ?)',
                    (name, code, study_program_id)
                )
                db.commit()
                flash(f'Kolegij "{name}" je dodan.', 'success')
                return redirect(url_for('course.index'))
            except db.IntegrityError:
                flash(f'Šifra kolegija "{code}" već postoji za ovaj studijski program.', 'danger')
    return render_template('course/form.html', programs=programs)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
def edit(id):
    db = get_db()
    course = db.execute('SELECT * FROM course WHERE id = ?', (id,)).fetchone()
    if course is None:
        flash('Kolegij nije pronađen.', 'danger')
        return redirect(url_for('course.index'))

    programs = db.execute('SELECT * FROM study_program ORDER BY name, element').fetchall()

    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip()
        study_program_id = request.form.get('study_program_id', type=int)
        if not name or not code or not study_program_id:
            flash('Sva polja su obavezna.', 'danger')
        else:
            try:
                db.execute(
                    'UPDATE course SET name = ?, code = ?, study_program_id = ? WHERE id = ?',
                    (name, code, study_program_id, id)
                )
                db.commit()
                flash('Kolegij je ažuriran.', 'success')
                return redirect(url_for('course.index'))
            except db.IntegrityError:
                flash(f'Šifra kolegija "{code}" već postoji za ovaj studijski program.', 'danger')
    return render_template('course/form.html', course=course, programs=programs)


@bp.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    db = get_db()
    db.execute('DELETE FROM course WHERE id = ?', (id,))
    db.commit()
    flash('Kolegij je obrisan.', 'success')
    return redirect(url_for('course.index'))


@bp.route('/api/by-program/<int:program_id>')
def api_by_program(program_id):
    """Vrati kolegije za zadani studijski program (AJAX)."""
    db = get_db()
    courses = db.execute(
        'SELECT id, name, code FROM course WHERE study_program_id = ? ORDER BY name',
        (program_id,)
    ).fetchall()
    return jsonify([{'id': c['id'], 'name': c['name'], 'code': c['code']} for c in courses])


@bp.route('/import', methods=['POST'])
def import_bulk():
    from openpyxl import load_workbook
    import io

    if 'file' not in request.files:
        flash('Datoteka nije odabrana.', 'danger')
        return redirect(url_for('course.index'))

    file = request.files['file']
    if file.filename == '':
        flash('Datoteka nije odabrana.', 'danger')
        return redirect(url_for('course.index'))

    study_program_id = request.form.get('study_program_id', type=int)
    if not study_program_id:
        flash('Odaberite studijski program za uvoz kolegija.', 'danger')
        return redirect(url_for('course.index'))

    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True)
        ws = wb.active
        db = get_db()
        added = 0
        skipped = 0
        duplicates = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            # Format: šifra, naziv kolegija
            cells = [str(c).strip() if c is not None else '' for c in row[:2]]
            code = cells[0]
            name = cells[1]

            if not code or not name:
                skipped += 1
                continue

            # Skip header row
            if code.lower() in ('šifra', 'sifra', 'code') and name.lower() in ('naziv', 'name', 'naziv kolegija'):
                continue

            try:
                db.execute(
                    'INSERT INTO course (name, code, study_program_id) VALUES (?, ?, ?)',
                    (name, code, study_program_id)
                )
                added += 1
            except db.IntegrityError:
                duplicates += 1

        db.commit()
        wb.close()
        msg = f'Uvezeno {added} kolegija.'
        if duplicates:
            msg += f' Preskočeno {duplicates} duplikata.'
        if skipped:
            msg += f' Preskočeno {skipped} neispravnih redaka.'
        flash(msg, 'success')
    except Exception as e:
        flash(f'Greška pri uvozu: {e}', 'danger')

    return redirect(url_for('course.index'))
