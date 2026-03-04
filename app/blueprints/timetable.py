import io
from datetime import date, datetime
from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash, jsonify
from app.auth import login_required, is_admin as check_admin
from app.db import get_db
from app.audit import log_audit
from app.models import (
    DAYS, WEEK_TYPES, SEMESTER_TYPES, STUDY_MODES,
    get_schedule_entries, build_timetable_grid, compute_day_columns, build_cell_info,
    build_program_colors, build_day_dates, get_display_days, get_week_dates, get_week_date_range,
    get_time_slots, get_merged_time_slots, check_conflicts, weeks_overlap, group_entries_by_week,
    sort_classrooms, sort_professors, sort_programs,
)

bp = Blueprint('timetable', __name__)


def _parse_date(date_str):
    """Parse dd.mm.YYYY. or dd.mm.YYYY to ISO YYYY-MM-DD."""
    date_str = date_str.strip().rstrip('.')
    try:
        return datetime.strptime(date_str, '%d.%m.%Y').strftime('%Y-%m-%d')
    except ValueError:
        return None


def _format_date(iso_date):
    """Format ISO YYYY-MM-DD to dd.mm.YYYY."""
    try:
        return datetime.strptime(iso_date, '%Y-%m-%d').strftime('%d.%m.%Y.')
    except (ValueError, TypeError):
        return iso_date or ''


def _get_default_academic_year_id():
    """Vrati ID zadane akademske godine ili None."""
    db = get_db()
    row = db.execute('SELECT id FROM academic_year WHERE is_default = 1').fetchone()
    return row['id'] if row else None


def _get_default_semester_type():
    """Vrati zadani tip semestra iz zadane akademske godine ili None."""
    db = get_db()
    row = db.execute('SELECT default_semester_type FROM academic_year WHERE is_default = 1').fetchone()
    return row['default_semester_type'] if row and row['default_semester_type'] else None


def get_day_statuses(academic_year_id, day_dates=None):
    """Load day statuses for an academic year as {day_of_week: {status, note}}.

    If day_dates is provided ({day_num: 'dd.mm.yyyy.'}), also loads date-specific
    statuses from day_status_date and merges them (date-specific takes priority).
    """
    if not academic_year_id:
        return {}
    db = get_db()

    # 1. Day-of-week statuses (existing behavior)
    rows = db.execute(
        'SELECT day_of_week, status, note FROM day_status WHERE academic_year_id = ?',
        (academic_year_id,)
    ).fetchall()
    result = {r['day_of_week']: {'status': r['status'], 'note': r['note']} for r in rows}

    # 2. Date-specific statuses (new)
    if day_dates:
        from datetime import datetime
        iso_pairs = []  # [(day_num, 'YYYY-MM-DD'), ...]
        for day_num, display_date in day_dates.items():
            for single_date in str(display_date).split(', '):
                single_date = single_date.strip().rstrip('.')
                try:
                    d = datetime.strptime(single_date, '%d.%m.%Y')
                    iso_pairs.append((day_num, d.strftime('%Y-%m-%d')))
                except ValueError:
                    pass

        if iso_pairs:
            all_iso = list(set(iso for _, iso in iso_pairs))
            placeholders = ','.join('?' * len(all_iso))
            date_rows = db.execute(
                f'SELECT date, status, note FROM day_status_date '
                f'WHERE academic_year_id = ? AND date IN ({placeholders})',
                [academic_year_id] + all_iso
            ).fetchall()
            date_map = {r['date']: {'status': r['status'], 'note': r['note']} for r in date_rows}

            for day_num, iso_date in iso_pairs:
                if iso_date in date_map:
                    result[day_num] = {**date_map[iso_date], '_date': iso_date}

    return result


def get_filter_options():
    db = get_db()
    if check_admin():
        programs = db.execute('SELECT * FROM study_program').fetchall()
    else:
        programs = db.execute('SELECT * FROM study_program WHERE is_service = 0').fetchall()
    return {
        'academic_years': db.execute('SELECT * FROM academic_year ORDER BY name DESC').fetchall(),
        'study_programs': sort_programs(programs),
        'professors': sort_professors(db.execute('SELECT * FROM professor').fetchall()),
        'classrooms': sort_classrooms(db.execute('SELECT * FROM classroom').fetchall()),
        'week_types': WEEK_TYPES,
        'semester_types': SEMESTER_TYPES,
        'study_modes': STUDY_MODES,
        'days': DAYS,
    }


def _apply_study_mode_context(filters):
    """Izračunaj display_days i day_dates iz study_mode i schedule_date filtera."""
    study_mode = filters.get('study_mode')
    schedule_date_raw = filters.get('schedule_date')
    display_days = get_display_days(study_mode)
    day_dates = {}

    # Resolve schedule_date: accept both dd.mm.YYYY. and YYYY-MM-DD
    schedule_date_iso = None
    if schedule_date_raw:
        if '.' in schedule_date_raw:
            schedule_date_iso = _parse_date(schedule_date_raw)
        else:
            schedule_date_iso = schedule_date_raw
            # Convert ISO to display format for template
            filters['schedule_date'] = _format_date(schedule_date_raw)

    if study_mode == 'izvanredni':
        if schedule_date_iso:
            # Specifični datum — prikaži samo taj tjedan
            day_dates = get_week_dates(schedule_date_iso, study_mode)
            date_from, date_to = get_week_date_range(schedule_date_iso, study_mode)
            filters['date_from'] = date_from
            filters['date_to'] = date_to
        # Bez datuma — dohvati sve, grupira se po tjednima u viewu

    return display_days, day_dates


def _build_per_week(entries, display_days, time_slots, study_mode, academic_year_id=None):
    """Grupiraj izvanredne unose po tjednima i izgradi grid za svaki tjedan."""
    weeks = group_entries_by_week(entries, study_mode)
    per_week = []
    for week in weeks:
        w_entries = week['entries']
        w_day_cols, w_entry_tracks, w_week_splits = compute_day_columns(w_entries, display_days)
        w_grid = build_timetable_grid(w_entries, display_days, time_slots)
        w_cell_info = build_cell_info(w_grid, time_slots, display_days, w_day_cols, w_entry_tracks, w_week_splits)
        w_program_colors = build_program_colors(w_entries)
        w_day_statuses = get_day_statuses(academic_year_id, week['day_dates']) if academic_year_id else {}
        per_week.append({
            'label': week['label'],
            'day_dates': week['day_dates'],
            'grid': w_grid,
            'cell_info': w_cell_info,
            'program_colors': w_program_colors,
            'day_columns': w_day_cols,
            'week_split_days': w_week_splits,
            'day_statuses': w_day_statuses,
        })
    return per_week


def _build_title_and_filters(view_type):
    """Build title and filters from request args based on view type."""
    filters = {}
    title = 'Raspored'
    db = get_db()

    if view_type == 'program':
        filters = {
            'academic_year_id': request.args.get('academic_year_id', type=int),
            'study_program_id': request.args.get('study_program_id', type=int),
            'semester_type': request.args.get('semester_type'),
            'semester_number': request.args.get('semester_number', type=int),
            'week_type': request.args.get('week_type'),
            'schedule_date': request.args.get('schedule_date'),
        }
        if filters.get('study_program_id'):
            prog = db.execute('SELECT * FROM study_program WHERE id = ?',
                              (filters['study_program_id'],)).fetchone()
            if prog:
                filters['study_mode'] = prog['study_mode']
                sem_num = filters.get('semester_number')
                sem_type = filters.get('semester_type', '')
                title = f"Raspored - {prog['name']} ({prog['study_mode'].capitalize()})"
                if sem_num:
                    title += f" - {sem_num}. semestar"
                if sem_type:
                    title += f" ({sem_type})"

    elif view_type == 'classroom':
        filters = {
            'academic_year_id': request.args.get('academic_year_id', type=int),
            'classroom_id': request.args.get('classroom_id', type=int),
            'week_type': request.args.get('week_type'),
            'study_mode': request.args.get('study_mode') or None,
            'schedule_date': request.args.get('schedule_date'),
        }
        if filters.get('classroom_id'):
            room = db.execute('SELECT * FROM classroom WHERE id = ?',
                              (filters['classroom_id'],)).fetchone()
            if room:
                title = f"Raspored - Učionica {room['name']}"

    elif view_type == 'professor':
        filters = {
            'academic_year_id': request.args.get('academic_year_id', type=int),
            'professor_id': request.args.get('professor_id', type=int),
            'week_type': request.args.get('week_type'),
            'study_mode': request.args.get('study_mode') or None,
            'schedule_date': request.args.get('schedule_date'),
        }
        if filters.get('professor_id'):
            prof = db.execute('SELECT * FROM professor WHERE id = ?',
                              (filters['professor_id'],)).fetchone()
            if prof:
                title = f"Raspored - {prof['title']} {prof['first_name']} {prof['last_name']}".strip()

    if 'academic_year_id' not in request.args and not filters.get('academic_year_id'):
        filters['academic_year_id'] = _get_default_academic_year_id()

    if filters.get('study_mode') == 'izvanredni':
        title += ' - Izvanredni'

    if filters.get('academic_year_id'):
        ay = db.execute('SELECT * FROM academic_year WHERE id = ?',
                        (filters['academic_year_id'],)).fetchone()
        if ay:
            title += f" ({ay['name']})"

    return title, filters


@bp.route('/program')
def by_program():
    db = get_db()
    filters = {
        'academic_year_id': request.args.get('academic_year_id', type=int),
        'study_program_id': request.args.get('study_program_id', type=int),
        'semester_type': request.args.get('semester_type'),
        'semester_number': request.args.get('semester_number', type=int),
        'week_type': request.args.get('week_type'),
        'schedule_date': request.args.get('schedule_date'),
    }
    if 'academic_year_id' not in request.args:
        filters['academic_year_id'] = _get_default_academic_year_id()
    if 'semester_type' not in request.args:
        default_st = _get_default_semester_type()
        if default_st == 'ispitni':
            return redirect(url_for('exam_timetable.index'))
        filters['semester_type'] = default_st

    # Determine study_mode from selected program
    study_mode = None
    prog = None
    if filters.get('study_program_id'):
        prog = db.execute('SELECT * FROM study_program WHERE id = ?',
                          (filters['study_program_id'],)).fetchone()
        if prog:
            study_mode = prog['study_mode']
    filters['study_mode'] = study_mode

    display_days, day_dates = _apply_study_mode_context(filters)

    time_slots = get_time_slots(filters.get('study_mode'), program=prog)
    if not check_admin():
        filters['published_only'] = True
    filters['include_service'] = True
    entries = get_schedule_entries(filters) if (filters.get('study_mode') and filters.get('semester_type')) else []
    if not day_dates and entries:
        day_dates = build_day_dates(entries, display_days)
    day_cols, entry_tracks, week_splits = compute_day_columns(entries, display_days)
    grid = build_timetable_grid(entries, display_days, time_slots)
    cell_info = build_cell_info(grid, time_slots, display_days, day_cols, entry_tracks, week_splits)
    program_colors = build_program_colors(entries)

    day_statuses = get_day_statuses(filters.get('academic_year_id'), day_dates)

    # Build print title
    print_title = ''
    if filters.get('study_program_id'):
        prog = db.execute('SELECT name, element, study_mode FROM study_program WHERE id = ?',
                          (filters['study_program_id'],)).fetchone()
        if prog:
            pname = f"{prog['name']} ({prog['study_mode'].capitalize()})"
            parts = [pname]
            if filters.get('semester_number'):
                sem = f"{filters['semester_number']}. semestar"
                if filters.get('semester_type'):
                    sem += f" ({filters['semester_type']})"
                parts.append(sem)
            elif filters.get('semester_type'):
                parts.append(f"{filters['semester_type'].capitalize()} semestar")
            print_title = ' - '.join(parts)

    # Per-semester grids when showing all semesters (no specific semester_number)
    # Skip for izvanredni — per_week is used instead
    per_semester = []
    if entries and not filters.get('semester_number') and study_mode != 'izvanredni':
        from collections import defaultdict
        grouped = defaultdict(list)
        service_entries = []
        for e in entries:
            if e['is_service']:
                service_entries.append(e)
            else:
                grouped[e['semester_number']].append(e)
        for sem_num in sorted(grouped.keys()):
            sem_entries = grouped[sem_num] + service_entries
            sem_day_cols, sem_entry_tracks, sem_week_splits = compute_day_columns(sem_entries, display_days)
            sem_grid = build_timetable_grid(sem_entries, display_days, time_slots)
            sem_cell_info = build_cell_info(sem_grid, time_slots, display_days, sem_day_cols, sem_entry_tracks, sem_week_splits)
            sem_program_colors = build_program_colors(sem_entries)
            per_semester.append({
                'number': sem_num,
                'grid': sem_grid,
                'cell_info': sem_cell_info,
                'program_colors': sem_program_colors,
                'day_columns': sem_day_cols,
                'week_split_days': sem_week_splits,
            })

    # Per-week grids for izvanredni without specific date
    per_week = []
    if entries and study_mode == 'izvanredni' and not filters.get('schedule_date'):
        per_week = _build_per_week(entries, display_days, time_slots, study_mode, filters.get('academic_year_id'))

    return render_template(
        'timetable/by_program.html',
        grid=grid, cell_info=cell_info, entries=entries, filters=filters,
        program_colors=program_colors, day_statuses=day_statuses, day_columns=day_cols,
        display_days=display_days, day_dates=day_dates, time_slots=time_slots,
        print_title=print_title, week_split_days=week_splits,
        per_semester=per_semester, per_week=per_week,
        **get_filter_options()
    )


@bp.route('/classroom')
def by_classroom():
    filters = {
        'academic_year_id': request.args.get('academic_year_id', type=int),
        'classroom_id': request.args.get('classroom_id', type=int),
        'semester_type': request.args.get('semester_type'),
        'week_type': request.args.get('week_type'),
        'study_mode': request.args.get('study_mode') or None,
        'schedule_date': request.args.get('schedule_date'),
    }
    if 'academic_year_id' not in request.args:
        filters['academic_year_id'] = _get_default_academic_year_id()
    if 'semester_type' not in request.args:
        default_st = _get_default_semester_type()
        if default_st == 'ispitni':
            return redirect(url_for('exam_timetable.index'))
        filters['semester_type'] = default_st

    display_days, day_dates = _apply_study_mode_context(filters)
    if not check_admin():
        filters['published_only'] = True

    entries = get_schedule_entries(filters) if (filters.get('study_mode') and filters.get('semester_type') and (filters.get('classroom_id') or filters.get('academic_year_id'))) else []
    time_slots = get_merged_time_slots(entries, filters.get('study_mode'))
    if not day_dates and entries:
        day_dates = build_day_dates(entries, display_days)
    day_cols, entry_tracks, week_splits = compute_day_columns(entries, display_days)
    grid = build_timetable_grid(entries, display_days, time_slots)
    cell_info = build_cell_info(grid, time_slots, display_days, day_cols, entry_tracks, week_splits)
    program_colors = build_program_colors(entries)

    day_statuses = get_day_statuses(filters.get('academic_year_id'), day_dates)

    # Build print title
    print_title = ''
    if filters.get('classroom_id'):
        db = get_db()
        room = db.execute('SELECT name FROM classroom WHERE id = ?',
                          (filters['classroom_id'],)).fetchone()
        if room:
            print_title = f"Učionica {room['name']}"
    elif entries:
        print_title = 'Sve učionice'

    study_mode = filters.get('study_mode')
    is_izvanredni = study_mode == 'izvanredni'

    # Per-classroom × per-week grids for izvanredni (all classrooms, no specific date)
    per_classroom_week = []
    if entries and is_izvanredni and not filters.get('schedule_date') and not filters.get('classroom_id'):
        from collections import defaultdict
        grouped = defaultdict(list)
        for e in entries:
            grouped[e['classroom_id']].append(e)
        for cid in sorted(grouped, key=lambda c: grouped[c][0]['classroom_name']):
            c_entries = grouped[cid]
            c_name = c_entries[0]['classroom_name']
            weeks = _build_per_week(c_entries, display_days, time_slots, 'izvanredni', filters.get('academic_year_id'))
            for week in weeks:
                per_classroom_week.append({
                    'label': f"Učionica {c_name} | {week['label']}",
                    'day_dates': week['day_dates'],
                    'grid': week['grid'],
                    'cell_info': week['cell_info'],
                    'program_colors': week['program_colors'],
                    'day_columns': week['day_columns'],
                    'week_split_days': week['week_split_days'],
                    'day_statuses': week['day_statuses'],
                })

    # Per-week grids for izvanredni with specific classroom (no date)
    per_week = []
    if entries and is_izvanredni and not filters.get('schedule_date') and filters.get('classroom_id'):
        per_week = _build_per_week(entries, display_days, time_slots, 'izvanredni', filters.get('academic_year_id'))

    # Per-classroom grids for redoviti (all classrooms)
    per_classroom = []
    if entries and not is_izvanredni and not filters.get('classroom_id'):
        from collections import defaultdict
        grouped = defaultdict(list)
        for e in entries:
            grouped[e['classroom_id']].append(e)
        for cid in sorted(grouped, key=lambda c: grouped[c][0]['classroom_name']):
            c_entries = grouped[cid]
            c_day_cols, c_entry_tracks, c_week_splits = compute_day_columns(c_entries, display_days)
            c_grid = build_timetable_grid(c_entries, display_days, time_slots)
            c_cell_info = build_cell_info(c_grid, time_slots, display_days, c_day_cols, c_entry_tracks, c_week_splits)
            c_program_colors = build_program_colors(c_entries)
            per_classroom.append({
                'name': c_entries[0]['classroom_name'],
                'grid': c_grid,
                'cell_info': c_cell_info,
                'program_colors': c_program_colors,
                'day_columns': c_day_cols,
                'week_split_days': c_week_splits,
            })

    return render_template(
        'timetable/by_classroom.html',
        grid=grid, cell_info=cell_info, entries=entries, filters=filters,
        program_colors=program_colors, day_statuses=day_statuses, day_columns=day_cols,
        display_days=display_days, day_dates=day_dates, time_slots=time_slots,
        per_classroom=per_classroom, print_title=print_title,
        week_split_days=week_splits, per_week=per_week,
        per_classroom_week=per_classroom_week,
        **get_filter_options()
    )


@bp.route('/professor')
def by_professor():
    filters = {
        'academic_year_id': request.args.get('academic_year_id', type=int),
        'professor_id': request.args.get('professor_id', type=int),
        'semester_type': request.args.get('semester_type'),
        'week_type': request.args.get('week_type'),
        'study_mode': request.args.get('study_mode') or None,
        'schedule_date': request.args.get('schedule_date'),
    }
    if 'academic_year_id' not in request.args:
        filters['academic_year_id'] = _get_default_academic_year_id()
    if 'semester_type' not in request.args:
        default_st = _get_default_semester_type()
        if default_st == 'ispitni':
            return redirect(url_for('exam_timetable.index'))
        filters['semester_type'] = default_st

    display_days, day_dates = _apply_study_mode_context(filters)
    if not check_admin():
        filters['published_only'] = True

    entries = get_schedule_entries(filters) if (filters.get('study_mode') and filters.get('semester_type') and filters.get('professor_id')) else []
    time_slots = get_merged_time_slots(entries, filters.get('study_mode'))
    if not day_dates and entries:
        day_dates = build_day_dates(entries, display_days)
    day_cols, entry_tracks, week_splits = compute_day_columns(entries, display_days)
    grid = build_timetable_grid(entries, display_days, time_slots)
    cell_info = build_cell_info(grid, time_slots, display_days, day_cols, entry_tracks, week_splits)
    program_colors = build_program_colors(entries)

    day_statuses = get_day_statuses(filters.get('academic_year_id'), day_dates)

    # Build print title
    print_title = ''
    if filters.get('professor_id'):
        db = get_db()
        prof = db.execute('SELECT title, first_name, last_name FROM professor WHERE id = ?',
                          (filters['professor_id'],)).fetchone()
        if prof:
            print_title = f"{prof['title']} {prof['first_name']} {prof['last_name']}".strip()

    # Per-week grids for izvanredni without specific date
    per_week = []
    if entries and filters.get('study_mode') == 'izvanredni' and not filters.get('schedule_date'):
        per_week = _build_per_week(entries, display_days, time_slots, 'izvanredni', filters.get('academic_year_id'))

    return render_template(
        'timetable/by_professor.html',
        grid=grid, cell_info=cell_info, entries=entries, filters=filters,
        program_colors=program_colors, day_statuses=day_statuses, day_columns=day_cols,
        display_days=display_days, day_dates=day_dates, time_slots=time_slots,
        print_title=print_title, week_split_days=week_splits, per_week=per_week,
        **get_filter_options()
    )


@bp.route('/conflicts')
@login_required
def conflicts():
    academic_year_id = request.args.get('academic_year_id', type=int)
    if not academic_year_id:
        academic_year_id = _get_default_academic_year_id()

    db = get_db()

    # Dohvati sve unose s konfliktom
    entries = db.execute('''
        SELECT se.*, c.name as course_name,
               p.first_name, p.last_name, p.title,
               cl.name as classroom_name,
               sp.name as program_name, sp.element as program_element, sp.study_mode
        FROM schedule_entry se
        JOIN course c ON se.course_id = c.id
        JOIN professor p ON se.professor_id = p.id
        JOIN classroom cl ON se.classroom_id = cl.id
        JOIN study_program sp ON se.study_program_id = sp.id
        WHERE se.has_conflict = 1 AND se.academic_year_id = ?
        ORDER BY se.day_of_week, se.start_time, se.classroom_id
    ''', (academic_year_id,)).fetchall()

    # Za svaki unos pronađi opis konflikta
    conflict_list = []
    for e in entries:
        entry_data = {
            'academic_year_id': e['academic_year_id'],
            'day_of_week': e['day_of_week'],
            'start_time': e['start_time'],
            'end_time': e['end_time'],
            'week_type': e['week_type'],
            'professor_id': e['professor_id'],
            'classroom_id': e['classroom_id'],
            'study_program_id': e['study_program_id'],
            'semester_number': e['semester_number'],
            'group_name': e['group_name'],
            'date': e['date'],
        }
        reasons = check_conflicts(entry_data, exclude_id=e['id'])
        prof_name = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
        prog_name = f"{e['program_name']} ({e['study_mode'].capitalize()})"
        conflict_list.append({
            'id': e['id'],
            'day_name': DAYS.get(e['day_of_week'], ''),
            'day_of_week': e['day_of_week'],
            'start_time': e['start_time'],
            'end_time': e['end_time'],
            'course_name': e['course_name'],
            'professor': prof_name,
            'classroom': e['classroom_name'],
            'program': prog_name,
            'semester_number': e['semester_number'],
            'week_type': e['week_type'],
            'group_name': e['group_name'],
            'reasons': reasons,
        })

    academic_years = db.execute(
        'SELECT * FROM academic_year ORDER BY name DESC'
    ).fetchall()

    return render_template(
        'timetable/conflicts.html',
        conflicts=conflict_list,
        academic_years=academic_years,
        selected_year=academic_year_id,
    )


@bp.route('/conflicts/excel')
@login_required
def export_conflicts_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from itertools import groupby as itertools_groupby
    from operator import itemgetter

    academic_year_id = request.args.get('academic_year_id', type=int)
    group_by = request.args.get('group_by', 'all')  # all, professor, classroom
    if not academic_year_id:
        academic_year_id = _get_default_academic_year_id()

    db = get_db()
    entries = db.execute('''
        SELECT se.*, c.name as course_name,
               p.first_name, p.last_name, p.title,
               cl.name as classroom_name,
               sp.name as program_name, sp.element as program_element, sp.study_mode
        FROM schedule_entry se
        JOIN course c ON se.course_id = c.id
        JOIN professor p ON se.professor_id = p.id
        JOIN classroom cl ON se.classroom_id = cl.id
        JOIN study_program sp ON se.study_program_id = sp.id
        WHERE se.has_conflict = 1 AND se.academic_year_id = ?
        ORDER BY se.day_of_week, se.start_time, se.classroom_id
    ''', (academic_year_id,)).fetchall()

    ay = db.execute('SELECT name FROM academic_year WHERE id = ?', (academic_year_id,)).fetchone()
    ay_name = ay['name'] if ay else ''

    # Build conflict data list
    conflict_list = []
    for e in entries:
        entry_data = {
            'academic_year_id': e['academic_year_id'],
            'day_of_week': e['day_of_week'],
            'start_time': e['start_time'],
            'end_time': e['end_time'],
            'week_type': e['week_type'],
            'professor_id': e['professor_id'],
            'classroom_id': e['classroom_id'],
            'study_program_id': e['study_program_id'],
            'semester_number': e['semester_number'],
            'group_name': e['group_name'],
            'date': e['date'],
        }
        reasons = check_conflicts(entry_data, exclude_id=e['id'])
        prof_name = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
        prog_name = f"{e['program_name']} ({e['study_mode'].capitalize()})"
        if e['semester_number']:
            prog_name += f" ({e['semester_number']}.sem"
            if e['group_name']:
                prog_name += f", gr.{e['group_name']}"
            prog_name += ")"

        day_name = DAYS.get(e['day_of_week'], '')
        time_str = f"{e['start_time']}-{e['end_time']}"
        if e['week_type'] != 'kontinuirano':
            time_str += f" [{e['week_type']}]"

        conflict_list.append({
            'day_name': day_name, 'time_str': time_str,
            'course_name': e['course_name'], 'professor': prof_name,
            'classroom': e['classroom_name'], 'program': prog_name,
            'reasons': reasons,
        })

    wb = Workbook()
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    group_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    group_fill_prof = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    group_fill_room = PatternFill(start_color='0DCAF0', end_color='0DCAF0', fill_type='solid')
    cell_font = Font(name='Arial', size=9)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _write_conflict_sheet(ws, items, title, headers, widths, get_values):
        """Write a conflict table to a worksheet."""
        merge_end = chr(64 + len(headers))
        ws.merge_cells(f'A1:{merge_end}1')
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name='Arial', bold=True, size=14, color='DC3545')
        title_cell.alignment = Alignment(horizontal='center')

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
            ws.column_dimensions[chr(64 + col)].width = w

        row = 4
        for item in items:
            values = get_values(item)
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = cell_font
                c.alignment = left_align if col >= 3 else center_align
                c.border = thin_border
            reason_lines = len(item['reasons']) if item['reasons'] else 1
            ws.row_dimensions[row].height = max(20, reason_lines * 15)
            row += 1
        return row

    def _write_grouped_sheet(ws, items, group_key, title, headers, widths, get_values, group_fill):
        """Write grouped conflict tables to a worksheet."""
        merge_end = chr(64 + len(headers))
        ws.merge_cells(f'A1:{merge_end}1')
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name='Arial', bold=True, size=14, color='DC3545')
        title_cell.alignment = Alignment(horizontal='center')

        sorted_items = sorted(items, key=itemgetter(group_key))
        row = 3
        for group_name, group_items in itertools_groupby(sorted_items, key=itemgetter(group_key)):
            group_items = list(group_items)
            # Group header row
            ws.merge_cells(f'A{row}:{merge_end}{row}')
            gc = ws.cell(row=row, column=1, value=f'{group_name} ({len(group_items)})')
            gc.font = group_font
            gc.fill = group_fill
            gc.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1

            # Column headers
            for col, (h, w) in enumerate(zip(headers, widths), 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center_align
                c.border = thin_border
                ws.column_dimensions[chr(64 + col)].width = w
            row += 1

            # Data rows
            for item in group_items:
                values = get_values(item)
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = cell_font
                    c.alignment = left_align if col >= 3 else center_align
                    c.border = thin_border
                reason_lines = len(item['reasons']) if item['reasons'] else 1
                ws.row_dimensions[row].height = max(20, reason_lines * 15)
                row += 1
            row += 1  # Empty row between groups

    def _reason_text(item):
        return '\n'.join(item['reasons']) if item['reasons'] else 'Konflikt je možda riješen'

    if group_by == 'professor':
        ws = wb.active
        ws.title = 'Po profesoru'
        _write_grouped_sheet(
            ws, conflict_list, 'professor',
            f'Konflikti po profesoru — {ay_name}',
            ['Dan', 'Vrijeme', 'Kolegij', 'Učionica', 'Program', 'Opis konflikta'],
            [14, 16, 28, 14, 30, 50],
            lambda item: [item['day_name'], item['time_str'], item['course_name'],
                          item['classroom'], item['program'], _reason_text(item)],
            group_fill_prof,
        )
        suffix = '_PROFESORI'
    elif group_by == 'classroom':
        ws = wb.active
        ws.title = 'Po učionici'
        _write_grouped_sheet(
            ws, conflict_list, 'classroom',
            f'Konflikti po učionici — {ay_name}',
            ['Dan', 'Vrijeme', 'Kolegij', 'Profesor', 'Program', 'Opis konflikta'],
            [14, 16, 28, 28, 30, 50],
            lambda item: [item['day_name'], item['time_str'], item['course_name'],
                          item['professor'], item['program'], _reason_text(item)],
            group_fill_room,
        )
        suffix = '_UCIONICE'
    else:
        ws = wb.active
        ws.title = 'Konflikti'
        _write_conflict_sheet(
            ws, conflict_list,
            f'Konflikti u rasporedu — {ay_name}',
            ['Dan', 'Vrijeme', 'Kolegij', 'Profesor', 'Učionica', 'Program', 'Opis konflikta'],
            [14, 16, 28, 28, 14, 30, 50],
            lambda item: [item['day_name'], item['time_str'], item['course_name'],
                          item['professor'], item['classroom'], item['program'], _reason_text(item)],
        )
        suffix = ''

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"FOOZOS_KONFLIKTI{suffix}_{date.today().strftime('%d_%m_%Y')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/unpublished')
@login_required
def unpublished():
    academic_year_id = request.args.get('academic_year_id', type=int)
    if not academic_year_id:
        academic_year_id = _get_default_academic_year_id()

    db = get_db()

    entries = db.execute('''
        SELECT se.*, c.name as course_name,
               p.first_name, p.last_name, p.title,
               cl.name as classroom_name,
               sp.name as program_name, sp.element as program_element, sp.study_mode
        FROM schedule_entry se
        JOIN course c ON se.course_id = c.id
        JOIN professor p ON se.professor_id = p.id
        JOIN classroom cl ON se.classroom_id = cl.id
        JOIN study_program sp ON se.study_program_id = sp.id
        WHERE se.is_published = 0 AND se.academic_year_id = ?
        ORDER BY se.day_of_week, se.start_time, se.classroom_id
    ''', (academic_year_id,)).fetchall()

    entry_list = []
    for e in entries:
        prof_name = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
        prog_name = f"{e['program_name']} ({e['study_mode'].capitalize()})"
        entry_list.append({
            'id': e['id'],
            'day_name': DAYS.get(e['day_of_week'], ''),
            'day_of_week': e['day_of_week'],
            'start_time': e['start_time'],
            'end_time': e['end_time'],
            'course_name': e['course_name'],
            'professor': prof_name,
            'classroom': e['classroom_name'],
            'program': prog_name,
            'semester_number': e['semester_number'],
            'week_type': e['week_type'],
            'group_name': e['group_name'],
            'has_conflict': e['has_conflict'],
        })

    academic_years = db.execute(
        'SELECT * FROM academic_year ORDER BY name DESC'
    ).fetchall()

    return render_template(
        'timetable/unpublished.html',
        entries=entry_list,
        academic_years=academic_years,
        selected_year=academic_year_id,
    )


@bp.route('/publish-selected', methods=['POST'])
@login_required
def publish_selected():
    db = get_db()
    if 'publish_all' in request.form:
        count = db.execute('SELECT COUNT(*) FROM schedule_entry WHERE is_published = 0').fetchone()[0]
        if count > 0:
            db.execute('UPDATE schedule_entry SET is_published = 1 WHERE is_published = 0')
            log_audit('publish', 'schedule_entry', f'Objavljeno {count} stavki rasporeda')
            db.commit()
            flash(f'Uspješno objavljeno {count} stavki rasporeda.', 'success')
        else:
            flash('Nema neobjavljenih stavki.', 'info')
    else:
        entry_ids = request.form.getlist('entry_ids', type=int)
        if entry_ids:
            placeholders = ','.join('?' * len(entry_ids))
            count = db.execute(
                f'UPDATE schedule_entry SET is_published = 1 WHERE id IN ({placeholders}) AND is_published = 0',
                entry_ids
            ).rowcount
            if count > 0:
                log_audit('publish', 'schedule_entry', f'Objavljeno {count} odabranih stavki rasporeda')
                db.commit()
                flash(f'Uspješno objavljeno {count} stavki rasporeda.', 'success')
            else:
                db.commit()
                flash('Odabrane stavke su već objavljene.', 'info')
        else:
            flash('Niste odabrali nijednu stavku.', 'warning')
    return redirect(url_for('timetable.unpublished'))


@bp.route('/api/free-classrooms')
@login_required
def api_free_classrooms():
    day_of_week = request.args.get('day_of_week', type=int)
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')
    week_type = request.args.get('week_type', 'kontinuirano')
    academic_year_id = request.args.get('academic_year_id', type=int)

    if not all([day_of_week, start_time, end_time, academic_year_id]):
        return jsonify({'classrooms': []})

    # Determine which week_types overlap with the given week_type
    if week_type == 'kontinuirano':
        overlapping_types = ['kontinuirano', '1. tjedan', '2. tjedan']
    elif week_type == '1. tjedan':
        overlapping_types = ['kontinuirano', '1. tjedan']
    else:
        overlapping_types = ['kontinuirano', '2. tjedan']

    placeholders = ','.join('?' * len(overlapping_types))

    db = get_db()
    occupied = db.execute(f'''
        SELECT DISTINCT classroom_id FROM schedule_entry
        WHERE academic_year_id = ?
          AND day_of_week = ?
          AND start_time < ?
          AND end_time > ?
          AND week_type IN ({placeholders})
    ''', [academic_year_id, day_of_week, end_time, start_time] + overlapping_types).fetchall()

    occupied_ids = {row['classroom_id'] for row in occupied}

    all_classrooms = db.execute(
        'SELECT id, name FROM classroom ORDER BY name'
    ).fetchall()

    free = [{'id': c['id'], 'name': c['name']}
            for c in all_classrooms if c['id'] not in occupied_ids]

    return jsonify({'classrooms': free})


@bp.route('/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    view_type = request.args.get('view', 'program')
    title, filters = _build_title_and_filters(view_type)
    display_days, day_dates = _apply_study_mode_context(filters)

    if view_type == 'program':
        filters['include_service'] = True

    if not check_admin():
        filters['published_only'] = True

    entries = get_schedule_entries(filters) if (filters.get('study_mode') and any(filters.values())) else []

    # Za by_program koristimo program-specifične slotove, inače merged
    if view_type == 'program' and filters.get('study_program_id'):
        _db = get_db()
        _prog = _db.execute('SELECT * FROM study_program WHERE id = ?',
                            (filters['study_program_id'],)).fetchone()
        time_slots = get_time_slots(filters.get('study_mode'), program=_prog)
    else:
        time_slots = get_merged_time_slots(entries, filters.get('study_mode'))

    if not day_dates and entries:
        day_dates = build_day_dates(entries, display_days)
    day_cols, entry_tracks, week_splits = compute_day_columns(entries, display_days)
    grid = build_timetable_grid(entries, display_days, time_slots)
    ci = build_cell_info(grid, time_slots, display_days, day_cols, entry_tracks, week_splits)
    program_colors = build_program_colors(entries)
    day_statuses = get_day_statuses(filters.get('academic_year_id'), day_dates)

    wb = Workbook()

    # Shared styles
    header_font = Font(name='Arial', bold=True, color='FFD600', size=10)
    header_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    time_fill = PatternFill(start_color='F0F2F5', end_color='F0F2F5', fill_type='solid')
    time_font = Font(name='Arial', bold=True, size=9)
    entry_font = Font(name='Arial', size=8)
    med_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    day_fills = {
        1: PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid'),
        2: PatternFill(start_color='43A047', end_color='43A047', fill_type='solid'),
        3: PatternFill(start_color='8E24AA', end_color='8E24AA', fill_type='solid'),
        4: PatternFill(start_color='8D6E63', end_color='8D6E63', fill_type='solid'),
        5: PatternFill(start_color='00ACC1', end_color='00ACC1', fill_type='solid'),
        6: PatternFill(start_color='5C6BC0', end_color='5C6BC0', fill_type='solid'),
    }
    status_fills = {
        'neradni': PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid'),
        'praznik': PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid'),
        'nenastavni': PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid'),
    }
    status_fonts = {
        'neradni': Font(name='Arial', bold=True, color='FFFFFF', size=10),
        'praznik': Font(name='Arial', bold=True, color='FFFFFF', size=10),
        'nenastavni': Font(name='Arial', bold=True, color='FFFFFF', size=10),
    }
    status_labels = {
        'neradni': 'Neradni dan',
        'praznik': 'Praznik',
        'nenastavni': 'Nenastavni dan',
    }
    day_off_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    conflict_fill = PatternFill(start_color='FF1744', end_color='FF1744', fill_type='solid')
    conflict_font = Font(name='Arial', size=8, color='FFFFFF')
    service_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    service_text_color = 'E65100'  # orange — matches web badge
    service_inline_font = InlineFont(b=True, color=service_text_color)

    def _col_letter(col_idx):
        result = ''
        while col_idx > 0:
            col_idx, rem = divmod(col_idx - 1, 26)
            result = chr(65 + rem) + result
        return result

    note_inline_font = InlineFont(b=True, color='FF0000')

    def _format_entry(e, vt):
        parts = [e['course_name'], f"{e['start_time']}-{e['end_time']}"]
        if vt != 'professor':
            prof = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
            parts.append(prof)
        if vt != 'classroom':
            parts.append(e['classroom_name'])
        if e['is_service']:
            parts.append(f"[Servis] {e['program_name']}")
        elif vt in ('classroom', 'professor'):
            pname = f"{e['program_name']} ({e['study_mode'].capitalize()})"
            parts.append(f"{pname} ({e['semester_number']}.sem)")
        if e['group_name']:
            parts.append(f"Grupa: {e['group_name']}")
        if e['module_name']:
            parts.append(f"Modul: {e['module_name']}")
        if e['week_type'] != 'kontinuirano':
            parts.append(f"[{e['week_type']}]")
        if e['note']:
            parts.append(f"* {e['note']}")
        return '\n'.join(parts)

    def _format_entry_rich(entries_list, vt, pc_map=None):
        """Build CellRichText with notes in red+bold, service entries in orange, or plain string."""
        has_any_note = any(e['note'] for e in entries_list)
        has_service = any(e['is_service'] for e in entries_list)
        program_ids = set(e['study_program_id'] for e in entries_list)
        multi_program = pc_map and len(program_ids) > 1

        if not has_any_note and not multi_program and not has_service:
            if len(entries_list) == 1:
                return _format_entry(entries_list[0], vt)
            return '\n---\n'.join(_format_entry(e, vt) for e in entries_list)

        segments = []
        for i, e in enumerate(entries_list):
            if i > 0:
                segments.append('\n---\n')
            parts = [e['course_name'], f"{e['start_time']}-{e['end_time']}"]
            if vt != 'professor':
                prof = f"{e['title']} {e['first_name']} {e['last_name']}".strip()
                parts.append(prof)
            if vt != 'classroom':
                parts.append(e['classroom_name'])
            if e['is_service']:
                parts.append(f"[Servis] {e['program_name']}")
            elif vt in ('classroom', 'professor'):
                pname = f"{e['program_name']} ({e['study_mode'].capitalize()})"
                parts.append(f"{pname} ({e['semester_number']}.sem)")
            if e['group_name']:
                parts.append(f"Grupa: {e['group_name']}")
            if e['module_name']:
                parts.append(f"Modul: {e['module_name']}")
            if e['week_type'] != 'kontinuirano':
                parts.append(f"[{e['week_type']}]")
            main_text = '\n'.join(parts)
            # Service entries always get orange text; multi-program gets palette color
            if e['is_service']:
                segments.append(TextBlock(service_inline_font, main_text))
            elif multi_program and pc_map:
                pc = pc_map.get(e['study_program_id'])
                if pc:
                    segments.append(TextBlock(InlineFont(color=pc['text'].lstrip('#')), main_text))
                else:
                    segments.append(main_text)
            else:
                segments.append(main_text)
            if e['note']:
                segments.append(TextBlock(note_inline_font, f"\n* {e['note']}"))
        return CellRichText(*segments)

    def _write_sheet(ws, sheet_title, sheet_day_cols, sheet_ci, sheet_program_colors, vt, sheet_week_splits=None, sheet_time_slots=None, sheet_day_dates=None, sheet_day_statuses=None):
        """Write a timetable grid to the given worksheet."""
        if sheet_week_splits is None:
            sheet_week_splits = set()
        if sheet_day_dates is None:
            sheet_day_dates = day_dates
        if sheet_time_slots is None:
            sheet_time_slots = time_slots
        if sheet_day_statuses is None:
            sheet_day_statuses = day_statuses

        # Column mapping
        day_col_start = {}
        col_cursor = 2
        for day_num in display_days:
            day_col_start[day_num] = col_cursor
            col_cursor += sheet_day_cols.get(day_num, 1)
        total_cols = col_cursor - 1

        has_splits = len(sheet_week_splits) > 0

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_cols, 2))
        title_cell = ws.cell(row=1, column=1, value=sheet_title)
        title_cell.font = Font(name='Arial', bold=True, size=14, color='2C5F8A')
        title_cell.alignment = Alignment(horizontal='center')

        # Header row
        header_row = 3
        if has_splits:
            ws.merge_cells(start_row=header_row, start_column=1,
                           end_row=header_row + 1, end_column=1)
            ws.cell(row=header_row + 1, column=1).border = med_border
        c = ws.cell(row=header_row, column=1, value='VRIJEME')
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = med_border
        ws.column_dimensions['A'].width = 16

        sub_header_font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
        sub_header_fill = PatternFill(start_color='455A64', end_color='455A64', fill_type='solid')

        for day_num, day_name in display_days.items():
            header_label = day_name.upper()
            if sheet_day_dates.get(day_num):
                header_label += f"\n{sheet_day_dates[day_num]}"
            ds = sheet_day_statuses.get(day_num)
            if ds:
                status_text = status_labels.get(ds['status'], ds['status'])
                if ds.get('note'):
                    status_text += f" - {ds['note']}"
                header_label += f"\n{status_text}"
            start_col = day_col_start[day_num]
            span = sheet_day_cols.get(day_num, 1)
            is_week_split = day_num in sheet_week_splits

            if span > 1:
                ws.merge_cells(start_row=header_row, start_column=start_col,
                               end_row=header_row, end_column=start_col + span - 1)
                for mc in range(start_col, start_col + span):
                    ws.cell(row=header_row, column=mc).border = med_border
            if has_splits and not is_week_split:
                ws.merge_cells(start_row=header_row, start_column=start_col,
                               end_row=header_row + 1, end_column=start_col + span - 1)
                for mc in range(start_col, start_col + span):
                    ws.cell(row=header_row + 1, column=mc).border = med_border

            c = ws.cell(row=header_row, column=start_col, value=header_label)
            if ds and ds['status'] in status_fills:
                c.font = status_fonts[ds['status']]
                c.fill = status_fills[ds['status']]
            else:
                c.font = header_font
                c.fill = day_fills.get(day_num, header_fill)
            c.alignment = center_align
            c.border = med_border
            is_date_view = bool(sheet_day_dates)
            for sc in range(start_col, start_col + span):
                base_w = 52 if is_date_view else 22
                ws.column_dimensions[_col_letter(sc)].width = max(18, base_w // span + 4)
                if sc != start_col:
                    hc = ws.cell(row=header_row, column=sc)
                    if ds and ds['status'] in status_fills:
                        hc.fill = status_fills[ds['status']]
                    else:
                        hc.fill = day_fills.get(day_num, header_fill)
                    hc.border = med_border

            # Sub-header row for week-split days
            if has_splits and is_week_split:
                for i, label in enumerate(['1. tj', '2. tj']):
                    sc = start_col + i
                    hc = ws.cell(row=header_row + 1, column=sc, value=label)
                    hc.font = sub_header_font
                    hc.fill = sub_header_fill
                    hc.alignment = center_align
                    hc.border = med_border

        # Data rows
        base_row = header_row + (2 if has_splits else 1)
        row_max_lines = {}  # Track max text lines per row for dynamic height
        for ts_idx, ts in enumerate(sheet_time_slots):
            r = base_row + ts_idx
            row_max_lines[r] = 0
            time_cell = ws.cell(row=r, column=1, value=ts)
            time_cell.font = time_font
            time_cell.fill = time_fill
            time_cell.alignment = center_align
            time_cell.border = med_border

            for day_num in display_days:
                tracks = sheet_ci[ts][day_num]
                start_col = day_col_start[day_num]
                is_day_off = day_num in sheet_day_statuses

                for track_idx, info in enumerate(tracks):
                    sc = start_col + track_idx

                    if info['skip']:
                        continue

                    # Skip cells that are already part of a merged range
                    if isinstance(ws.cell(row=r, column=sc), MergedCell):
                        continue

                    rowspan = info['rowspan']
                    colspan = info.get('colspan', 1)

                    if info['entries']:
                        cell_text = _format_entry_rich(info['entries'], vt, sheet_program_colors)
                        # Track text lines for dynamic row height
                        plain = str(cell_text) if isinstance(cell_text, CellRichText) else cell_text
                        text_lines = plain.count('\n') + 1
                        effective_lines = text_lines // max(rowspan, 1)
                        if effective_lines > row_max_lines.get(r, 0):
                            row_max_lines[r] = effective_lines
                        has_conflict = any(e['has_conflict'] for e in info['entries'])
                        end_row = r + rowspan - 1 if rowspan > 1 else r
                        end_col = sc + colspan - 1 if colspan > 1 else sc
                        if rowspan > 1 or colspan > 1:
                            ws.merge_cells(start_row=r, start_column=sc,
                                           end_row=end_row, end_column=end_col)
                            # Apply borders to all cells AFTER merging
                            for mr in range(r, end_row + 1):
                                for mc in range(sc, end_col + 1):
                                    ws.cell(row=mr, column=mc).border = thin_border
                        c = ws.cell(row=r, column=sc, value=cell_text)
                        c.alignment = center_align
                        c.border = thin_border
                        all_service = all(e['is_service'] for e in info['entries'])
                        if has_conflict:
                            c.fill = conflict_fill
                            c.font = conflict_font
                        elif is_day_off:
                            c.fill = day_off_fill
                            c.font = entry_font
                        elif all_service:
                            # Service-only cells: orange background (matches web badge)
                            c.fill = service_fill
                            c.font = Font(name='Arial', size=8,
                                          color=service_text_color)
                        else:
                            # Use first non-service entry's color for cell background
                            first_entry = next(
                                (e for e in info['entries'] if not e['is_service']),
                                info['entries'][0]
                            )
                            pc = sheet_program_colors.get(first_entry['study_program_id'])
                            if pc:
                                c.fill = PatternFill(
                                    start_color=pc['bg'].lstrip('#'),
                                    end_color=pc['bg'].lstrip('#'),
                                    fill_type='solid'
                                )
                                c.font = Font(name='Arial', size=8,
                                              color=pc['text'].lstrip('#'))
                            else:
                                c.font = entry_font
                    else:
                        c = ws.cell(row=r, column=sc, value='')
                        c.border = thin_border
                        if is_day_off:
                            c.fill = day_off_fill

        # Set dynamic row heights based on content
        for r in range(base_row, base_row + len(sheet_time_slots)):
            max_lines = row_max_lines.get(r, 0)
            # Base height 60, add 13 per extra line beyond 5
            ws.row_dimensions[r].height = max(60, 60 + (max_lines - 5) * 13) if max_lines > 5 else 60

        # Final pass: ensure ALL data cells have borders (including merged cells)
        for r in range(base_row, base_row + len(sheet_time_slots)):
            for c in range(2, total_cols + 1):
                cell = ws.cell(row=r, column=c)
                if not cell.border or not cell.border.left.style:
                    cell.border = thin_border

    # Per-week sheets for izvanredni without specific date
    study_mode_excel = filters.get('study_mode')
    is_izvanredni_all = study_mode_excel == 'izvanredni' and not filters.get('schedule_date') and entries

    # Per-classroom × per-week sheets for izvanredni classroom view (all classrooms)
    if view_type == 'classroom' and is_izvanredni_all and not filters.get('classroom_id'):
        from collections import defaultdict
        grouped = defaultdict(list)
        for e in entries:
            grouped[e['classroom_id']].append(e)
        first = True
        for cid in sorted(grouped, key=lambda c: grouped[c][0]['classroom_name']):
            c_entries = grouped[cid]
            c_name = c_entries[0]['classroom_name']
            weeks = _build_per_week(c_entries, display_days, time_slots, 'izvanredni', filters.get('academic_year_id'))
            for week in weeks:
                sheet_label = f"Uč. {c_name} | {week['label']}"
                sheet_name = sheet_label[:31]
                if first:
                    w_ws = wb.active
                    w_ws.title = sheet_name
                    first = False
                else:
                    w_ws = wb.create_sheet(title=sheet_name)
                _write_sheet(w_ws, f"Učionica {c_name} | {week['label']}", week['day_columns'], week['cell_info'],
                             week['program_colors'], view_type, week['week_split_days'],
                             sheet_day_dates=week['day_dates'], sheet_day_statuses=week['day_statuses'])
    # Per-classroom sheets (all classrooms, redoviti) – skip combined main sheet
    elif view_type == 'classroom' and entries and not filters.get('classroom_id'):
        from collections import defaultdict
        grouped = defaultdict(list)
        for e in entries:
            grouped[e['classroom_id']].append(e)
        first = True
        for cid in sorted(grouped, key=lambda c: grouped[c][0]['classroom_name']):
            c_entries = grouped[cid]
            c_name = c_entries[0]['classroom_name']
            c_day_cols, c_entry_tracks, c_week_splits = compute_day_columns(c_entries, display_days)
            c_grid = build_timetable_grid(c_entries, display_days, time_slots)
            c_ci = build_cell_info(c_grid, time_slots, display_days, c_day_cols, c_entry_tracks, c_week_splits)
            c_program_colors = build_program_colors(c_entries)
            if first:
                c_ws = wb.active
                c_ws.title = c_name[:31]
                first = False
            else:
                c_ws = wb.create_sheet(title=c_name[:31])
            _write_sheet(c_ws, f"Učionica {c_name}", c_day_cols, c_ci, c_program_colors, 'classroom', c_week_splits)
    # Per-week sheets for izvanredni (program/professor/single classroom) without specific date
    elif is_izvanredni_all:
        per_week = _build_per_week(entries, display_days, time_slots, 'izvanredni', filters.get('academic_year_id'))
        first = True
        for week in per_week:
            sheet_name = week['label'][:31]
            if first:
                w_ws = wb.active
                w_ws.title = sheet_name
                first = False
            else:
                w_ws = wb.create_sheet(title=sheet_name)
            _write_sheet(w_ws, f"{title} - {week['label']}", week['day_columns'], week['cell_info'],
                         week['program_colors'], view_type, week['week_split_days'],
                         sheet_day_dates=week['day_dates'], sheet_day_statuses=week['day_statuses'])
    # Per-semester sheets (all semesters selected) – skip combined main sheet
    elif view_type == 'program' and entries and not filters.get('semester_number'):
        from collections import defaultdict
        grouped = defaultdict(list)
        service_entries = []
        for e in entries:
            if e['is_service']:
                service_entries.append(e)
            else:
                grouped[e['semester_number']].append(e)
        first = True
        for sem_num in sorted(grouped.keys()):
            sem_entries = grouped[sem_num] + service_entries
            s_day_cols, s_entry_tracks, s_week_splits = compute_day_columns(sem_entries, display_days)
            s_grid = build_timetable_grid(sem_entries, display_days, time_slots)
            s_ci = build_cell_info(s_grid, time_slots, display_days, s_day_cols, s_entry_tracks, s_week_splits)
            s_program_colors = build_program_colors(sem_entries)
            sheet_name = f"{sem_num}. semestar"
            if first:
                s_ws = wb.active
                s_ws.title = sheet_name[:31]
                first = False
            else:
                s_ws = wb.create_sheet(title=sheet_name[:31])
            _write_sheet(s_ws, f"{title} - {sheet_name}", s_day_cols, s_ci, s_program_colors, view_type, s_week_splits)
    else:
        # Single view – one main sheet
        ws = wb.active
        ws.title = 'Raspored'
        _write_sheet(ws, title, day_cols, ci, program_colors, view_type, week_splits)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    view_suffixes = {'program': 'STUDIJI', 'classroom': 'UCIONICE', 'professor': 'PROFESORI'}
    suffix = view_suffixes.get(view_type, 'RASPORED')
    if study_mode_excel == 'izvanredni':
        suffix += '_IZVANREDNI'
    filename = f"FOOZOS_RASPORED_{suffix}_{date.today().strftime('%d_%m_%Y')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
