from flask import Blueprint, render_template, request, redirect, url_for, flash
from app.db import get_db
from app.models import STUDY_MODES

bp = Blueprint('study_program', __name__)


@bp.route('/')
def index():
    db = get_db()
    programs = db.execute('SELECT * FROM study_program ORDER BY name').fetchall()
    return render_template('study_program/index.html', programs=programs)


@bp.route('/create', methods=['GET', 'POST'])
def create():
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip()
        study_mode = request.form.get('study_mode', 'redoviti')
        if not name or not code:
            flash('Sva polja su obavezna.', 'danger')
        else:
            db = get_db()
            try:
                db.execute(
                    'INSERT INTO study_program (name, code, study_mode) VALUES (?, ?, ?)',
                    (name, code, study_mode)
                )
                db.commit()
                flash(f'Studijski program "{name}" je dodan.', 'success')
                return redirect(url_for('study_program.index'))
            except db.IntegrityError:
                flash(f'Šifra "{code}" već postoji.', 'danger')
    return render_template('study_program/form.html', study_modes=STUDY_MODES)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
def edit(id):
    db = get_db()
    program = db.execute('SELECT * FROM study_program WHERE id = ?', (id,)).fetchone()
    if program is None:
        flash('Studijski program nije pronađen.', 'danger')
        return redirect(url_for('study_program.index'))

    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip()
        study_mode = request.form.get('study_mode', 'redoviti')
        if not name or not code:
            flash('Sva polja su obavezna.', 'danger')
        else:
            try:
                db.execute(
                    'UPDATE study_program SET name = ?, code = ?, study_mode = ? WHERE id = ?',
                    (name, code, study_mode, id)
                )
                db.commit()
                flash('Studijski program je ažuriran.', 'success')
                return redirect(url_for('study_program.index'))
            except db.IntegrityError:
                flash(f'Šifra "{code}" već postoji.', 'danger')
    return render_template('study_program/form.html', program=program, study_modes=STUDY_MODES)


@bp.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    db = get_db()
    db.execute('DELETE FROM study_program WHERE id = ?', (id,))
    db.commit()
    flash('Studijski program je obrisan.', 'success')
    return redirect(url_for('study_program.index'))


@bp.route('/import', methods=['POST'])
def import_bulk():
    from openpyxl import load_workbook
    import io

    if 'file' not in request.files:
        flash('Datoteka nije odabrana.', 'danger')
        return redirect(url_for('study_program.index'))

    file = request.files['file']
    if file.filename == '':
        flash('Datoteka nije odabrana.', 'danger')
        return redirect(url_for('study_program.index'))

    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True)
        ws = wb.active
        db = get_db()
        added = 0
        skipped = 0
        duplicates = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            # Format: šifra, naziv, način studiranja
            cells = [str(c).strip() if c is not None else '' for c in row[:3]]
            code = cells[0]
            name = cells[1]
            study_mode = cells[2].lower() if len(cells) > 2 and cells[2] else 'redoviti'

            if not code or not name:
                skipped += 1
                continue

            # Skip header row
            if code.lower() in ('šifra', 'sifra', 'code') and name.lower() in ('naziv', 'name'):
                continue

            if study_mode not in ('redoviti', 'izvanredni'):
                study_mode = 'redoviti'

            try:
                db.execute(
                    'INSERT INTO study_program (name, code, study_mode) VALUES (?, ?, ?)',
                    (name, code, study_mode)
                )
                added += 1
            except db.IntegrityError:
                duplicates += 1

        db.commit()
        wb.close()
        msg = f'Uvezeno {added} programa.'
        if duplicates:
            msg += f' Preskočeno {duplicates} duplikata.'
        if skipped:
            msg += f' Preskočeno {skipped} neispravnih redaka.'
        flash(msg, 'success')
    except Exception as e:
        flash(f'Greška pri uvozu: {e}', 'danger')

    return redirect(url_for('study_program.index'))
